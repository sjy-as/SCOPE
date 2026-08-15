"""
Ablation experiment execution framework: N LLMs × 3 datasets × multiple ablation variants.

Ablation models:
    SCOPE_wo_semantic_directory_few-shot     — SCOPE with few-shot routing instead of the semantic directory
    SCOPE_wo_semantic_directory_summary — SCOPE with summary profiles instead of the semantic directory
    SCOPE_wo_decomposition            — SCOPE without question decomposition
    SCOPE_wo_reflection          — SCOPE without reflection/fallback retry
    SCOPE_wo_consolidation       — SCOPE without runtime consolidation
    SCOPE_wo_semantic_guide_on_planning      — SCOPE without semantic guidance in operator planning
    SCOPE_wo_operator_planning            — SCOPE without operator planning

These compare against the full SCOPE (whose predictions are read from the main
experiment output directory). Run run_ex_main.py first to produce SCOPE predictions.

Layout (ablation experiment owns its own output tree):
    <SCOPE_ABLA_RESULT_ROOT or SCOPE_EVAL_DIR/results/ex_abla>/
        <dataset_slug>/<llm_slug>/
            SCOPE_wo_semantic_directory_few-shot/predictions.jsonl
            SCOPE_wo_semantic_directory_summary/predictions.jsonl
            eval_summary/summary.json   ← only the ablation variants are judged here
        tables/
            abla_tables.json + abla_tables.xlsx   (SCOPE row reads from
                                                   ex_main/.../eval_summary/summary.json)

Model lists for both experiments live in SCOPE_code/eval/models_config.py by default, or in SCOPE_EVAL_DIR when overridden.

Recommended invocation (run yourself):

    python3 run_ex_abla.py
    python3 run_ex_abla.py --only-llm deepseek-chat
    python3 run_ex_abla.py --tables-only
"""
from __future__ import annotations

import argparse
import json
import os
import shlex
import subprocess
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

sys.path.insert(0, str(Path(__file__).resolve().parent / "eval"))
from models_config import (  # noqa: E402
    ABLATION_EXPERIMENT_MODELS,
    ABLATION_TABLE_MODELS,
    MODEL_DISPLAY_NAMES,
    MODEL_EVAL_FLAGS,
    MODEL_TO_SOURCE,
    predictions_path,
)

try:
    from openpyxl import Workbook as _XlWorkbook
    from openpyxl.styles import Font as _XlFont, Alignment as _XlAlignment
    from openpyxl.styles import Border as _XlBorder, Side as _XlSide
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


# ---------------------------------------------------------------------------
# Static configuration
# ---------------------------------------------------------------------------

CODE_ROOT = Path(__file__).resolve().parent
NEW_MODEL_DIR = Path(os.getenv("SCOPE_NEW_MODEL_DIR", str(CODE_ROOT / "SCOPE"))).resolve()
def _resolve_ablation_baseline_dir() -> Path:
    explicit = os.getenv("SCOPE_ABLA_BASELINE_DIR")
    if explicit:
        return Path(explicit).resolve()

    default_dir = (CODE_ROOT / "wo_abla").resolve()
    legacy = os.getenv("SCOPE_BASELINE_DIR")
    if legacy:
        legacy_dir = Path(legacy).resolve()
        if (legacy_dir / "SCOPE_wo_decomposition" / "run.py").exists():
            return legacy_dir
        if (default_dir / "SCOPE_wo_decomposition" / "run.py").exists():
            print(f"[warn] SCOPE_BASELINE_DIR={legacy_dir} does not contain ablation wrappers; using {default_dir}", flush=True)
            return default_dir
        return legacy_dir

    return default_dir


BASELINE_DIR = _resolve_ablation_baseline_dir()
EVAL_DIR = Path(os.getenv("SCOPE_EVAL_DIR", str(CODE_ROOT / "eval"))).resolve()
QA_BENCH = Path(os.getenv("SCOPE_QA_BENCH", str(CODE_ROOT / "CMQA" / "qa_bench"))).resolve()
RESULT_ROOT = Path(os.getenv("SCOPE_ABLA_RESULT_ROOT", str(EVAL_DIR / "results" / "ex_abla"))).resolve()
# SCOPE (full version) predictions are read from the main experiment output
# directory — this runner only executes the ablation variants.
RESULT_ROOT_MAIN = Path(os.getenv("SCOPE_RESULT_ROOT", str(EVAL_DIR / "results" / "ex_main"))).resolve()
LOG_ROOT = RESULT_ROOT / "_logs"

DATASETS: Dict[str, Dict[str, Any]] = {
    "kg_doc":    {"file": QA_BENCH / "kg-doc-1154.jsonl",   "kb": "kg,doc",    "evaluator": "evaluate_answer_kg_doc.py"},
    "kg_table":  {"file": QA_BENCH / "kg-table-1147.jsonl", "kb": "kg,table",  "evaluator": "evaluate_answer_kg_table.py"},
    "table_doc": {"file": QA_BENCH / "table-doc-1120.jsonl","kb": "table,doc", "evaluator": "evaluate_answer_table_doc.py"},
}

DATASET_ORDER = ["kg_doc", "kg_table", "table_doc"]
# Models that get RUN by this script. Edit eval/models_config.py to extend.
ABLATION_MODELS = list(ABLATION_EXPERIMENT_MODELS)
LLM_ORDER = ["gpt-4o-mini", "gpt-4o", "qwen3-32b", "deepseek-v3", "deepseek-chat"]

# Display names: shared dict from eval/models_config.py, with "SCOPE" overridden
# to "SCOPE (Full)" so the ablation table reads naturally.
ABLATION_DISPLAY_NAMES: Dict[str, str] = {**MODEL_DISPLAY_NAMES, "SCOPE": "SCOPE (Full)"}

LLM_DISPLAY_NAMES: Dict[str, str] = {
    "gpt-4o-mini":                "GPT-4o-mini",
    "gpt-4o":                     "GPT-4o",
    "qwen3-32b":                  "Qwen3-32B",
    "deepseek-v3":                "DeepSeek-V3",
    "deepseek-chat":              "DeepSeek-Chat",
}

LLM_MODEL_IDS: Dict[str, str] = {
    "gpt-4o-mini": "gpt-4o-mini",
    "gpt-4o": "gpt-4o",
    "qwen3-32b": "Qwen3-32B",
    "deepseek-v3": "DeepSeek-V3",
    "deepseek-chat": "deepseek-chat",
    "chatanywhere_deepseek-chat": "deepseek-chat",
}


def _env_slug(name: str) -> str:
    return "".join(ch if ch.isalnum() else "_" for ch in name.upper())


def _llm_config(llm_slug: str) -> Dict[str, Any]:
    env_slug = _env_slug(llm_slug)
    base_url = os.getenv(f"{env_slug}_BASE_URL") or os.getenv("LLM_BASE_URL", "")
    api_key = os.getenv(f"{env_slug}_API_KEY") or os.getenv("LLM_API_KEY", "")
    model_name = os.getenv(f"{env_slug}_MODEL") or LLM_MODEL_IDS[llm_slug]
    return {
        "model": model_name,
        "base_url": base_url,
        "keys": {model: api_key for model in ABLATION_MODELS},
    }


# API configuration for ablation models. Set LLM_BASE_URL / LLM_API_KEY
# globally, or override one model family with e.g. DEEPSEEK_CHAT_API_KEY.
LLM_CONFIGS: Dict[str, Dict[str, Any]] = {
    llm_slug: _llm_config(llm_slug) for llm_slug in LLM_ORDER
}

# Evaluator (LLM-judge) configuration.
JUDGE_LLM_MODEL = os.getenv("JUDGE_LLM_MODEL", os.getenv("LLM_MODEL", "deepseek-chat"))
JUDGE_LLM_URL   = os.getenv("JUDGE_LLM_BASE_URL", os.getenv("LLM_BASE_URL", ""))
JUDGE_API_KEY   = os.getenv("JUDGE_LLM_API_KEY", os.getenv("LLM_API_KEY", ""))

PER_MODEL_WORKERS = 24
PER_LLM_WORKERS: Dict[str, int] = {
    "deepseek-chat": 24,
}


def workers_for(llm: str) -> int:
    return PER_LLM_WORKERS.get(llm, PER_MODEL_WORKERS)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_name(s: str) -> str:
    return s.replace("/", "_").replace(" ", "_")


def out_root(dataset: str, llm: str) -> Path:
    return RESULT_ROOT / dataset / _safe_name(llm)


def model_out_dir(dataset: str, llm: str, model: str) -> Path:
    return out_root(dataset, llm) / model


def log_path(dataset: str, llm: str, model: str) -> Path:
    LOG_ROOT.mkdir(parents=True, exist_ok=True)
    return LOG_ROOT / f"{dataset}__{_safe_name(llm)}__{model}.log"


def eval_log_path(dataset: str, llm: str) -> Path:
    LOG_ROOT.mkdir(parents=True, exist_ok=True)
    return LOG_ROOT / f"{dataset}__{_safe_name(llm)}__abla_eval.log"


def banner(msg: str) -> None:
    bar = "=" * 78
    print(f"\n{bar}\n{msg}\n{bar}", flush=True)


def file_has_content(p: Path) -> bool:
    return p.exists() and p.is_file() and p.stat().st_size > 0


def run_cmd(
    cmd: List[str],
    log_file: Path,
    cwd: Optional[Path] = None,
    env: Optional[Dict[str, str]] = None,
    timeout: Optional[int] = None,
    quiet: bool = False,
) -> Tuple[int, float]:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    print(f"[exec] {' '.join(shlex.quote(c) for c in cmd)}", flush=True)
    print(f"[exec] cwd={cwd}  log={log_file}", flush=True)
    t0 = time.time()
    with log_file.open("w", encoding="utf-8") as lf:
        lf.write(f"# CMD: {' '.join(shlex.quote(c) for c in cmd)}\n")
        if cwd:
            lf.write(f"# CWD: {cwd}\n")
        lf.flush()
        try:
            proc = subprocess.Popen(
                cmd, cwd=str(cwd) if cwd else None,
                env={**os.environ, **(env or {})},
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                bufsize=1, universal_newlines=True,
                encoding="utf-8", errors="replace",
            )
        except FileNotFoundError as e:
            lf.write(f"\n[ERROR] {e}\n")
            return 127, time.time() - t0

        assert proc.stdout is not None
        for line in proc.stdout:
            if not quiet:
                sys.stdout.write(line)
                sys.stdout.flush()
            lf.write(line)
        proc.wait(timeout=timeout)
    elapsed = time.time() - t0
    return proc.returncode or 0, elapsed


# ---------------------------------------------------------------------------
# Per-model command builders
# ---------------------------------------------------------------------------

def cmd_SCOPE_wo_semantic_directory_few_shot(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    """SCOPE ablation: AtomR source routing + no semantic-list content in parse/plan."""
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        "python3", str(BASELINE_DIR / "SCOPE_wo_semantic_directory_few-shot" / "run.py"),
        "--input", str(ds["file"]),
        "--gold",  str(ds["file"]),
        "--output-dir", str(out_dir),
        "--kb", ds["kb"],
        "--workers", str(PER_MODEL_WORKERS),
        "--api-key",  cfg["keys"]["SCOPE_wo_semantic_directory_few-shot"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
    ]
    return cmd, NEW_MODEL_DIR, {}


def cmd_SCOPE_wo_semantic_directory_summary(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    """SCOPE ablation: DeepSieve source routing + no semantic-list content in parse/plan."""
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        "python3", str(BASELINE_DIR / "SCOPE_wo_semantic_directory_summary" / "run.py"),
        "--input", str(ds["file"]),
        "--gold",  str(ds["file"]),
        "--output-dir", str(out_dir),
        "--kb", ds["kb"],
        "--workers", str(PER_MODEL_WORKERS),
        "--api-key",  cfg["keys"]["SCOPE_wo_semantic_directory_summary"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
    ]
    return cmd, NEW_MODEL_DIR, {}


def cmd_SCOPE_wo_decomposition(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    """SCOPE ablation: no question decomposition."""
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        "python3", str(BASELINE_DIR / "SCOPE_wo_decomposition" / "run.py"),
        "--input", str(ds["file"]),
        "--gold",  str(ds["file"]),
        "--output-dir", str(out_dir),
        "--kb", ds["kb"],
        "--workers", str(workers_for(llm)),
        "--api-key",  cfg["keys"]["SCOPE_wo_decomposition"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
    ]
    return cmd, NEW_MODEL_DIR, {}


def cmd_SCOPE_wo_reflection(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    """SCOPE ablation: no fallback-source reflection/retry."""
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        "python3", str(BASELINE_DIR / "SCOPE_wo_reflection" / "run.py"),
        "--input", str(ds["file"]),
        "--gold",  str(ds["file"]),
        "--output-dir", str(out_dir),
        "--kb", ds["kb"],
        "--workers", str(workers_for(llm)),
        "--api-key",  cfg["keys"]["SCOPE_wo_reflection"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
    ]
    return cmd, NEW_MODEL_DIR, {}


def cmd_SCOPE_wo_consolidation(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    """SCOPE ablation: no runtime consolidation memory."""
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        "python3", str(BASELINE_DIR / "SCOPE_wo_consolidation" / "run.py"),
        "--input", str(ds["file"]),
        "--gold",  str(ds["file"]),
        "--output-dir", str(out_dir),
        "--kb", ds["kb"],
        "--workers", str(workers_for(llm)),
        "--api-key",  cfg["keys"]["SCOPE_wo_consolidation"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
    ]
    return cmd, NEW_MODEL_DIR, {}


def cmd_SCOPE_wo_semantic_guide_on_planning(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    """SCOPE ablation: keep planner on, but remove semantic-list metadata from plan generation."""
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        "python3", str(BASELINE_DIR / "SCOPE_wo_semantic_guide_on_planning" / "run.py"),
        "--input", str(ds["file"]),
        "--gold",  str(ds["file"]),
        "--output-dir", str(out_dir),
        "--kb", ds["kb"],
        "--workers", str(workers_for(llm)),
        "--api-key",  cfg["keys"]["SCOPE_wo_semantic_guide_on_planning"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
    ]
    return cmd, NEW_MODEL_DIR, {}


def cmd_SCOPE_wo_operator_planning(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    """SCOPE ablation: no operator-tree planning (naive Search step)."""
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        "python3", str(BASELINE_DIR / "SCOPE_wo_operator_planning" / "run.py"),
        "--input", str(ds["file"]),
        "--gold",  str(ds["file"]),
        "--output-dir", str(out_dir),
        "--kb", ds["kb"],
        "--workers", str(workers_for(llm)),
        "--api-key",  cfg["keys"]["SCOPE_wo_operator_planning"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
    ]
    return cmd, NEW_MODEL_DIR, {}


MODEL_CMD_BUILDERS = {
    "SCOPE_wo_semantic_directory_few-shot":     cmd_SCOPE_wo_semantic_directory_few_shot,
    "SCOPE_wo_semantic_directory_summary": cmd_SCOPE_wo_semantic_directory_summary,
    "SCOPE_wo_decomposition":        cmd_SCOPE_wo_decomposition,
    "SCOPE_wo_reflection":      cmd_SCOPE_wo_reflection,
    "SCOPE_wo_consolidation":   cmd_SCOPE_wo_consolidation,
    "SCOPE_wo_semantic_guide_on_planning":  cmd_SCOPE_wo_semantic_guide_on_planning,
    "SCOPE_wo_operator_planning":        cmd_SCOPE_wo_operator_planning,
}


# ---------------------------------------------------------------------------
# Completion sentinels
# ---------------------------------------------------------------------------

@lru_cache(maxsize=None)
def dataset_size(dataset: str) -> int:
    gold = DATASETS[dataset]["file"]
    n = 0
    with gold.open("rb") as f:
        for _ in f:
            n += 1
    return n


def _count_jsonl_lines(p: Path) -> int:
    if not p.exists() or not p.is_file():
        return 0
    n = 0
    with p.open("rb") as f:
        for _ in f:
            n += 1
    return n


def model_output_progress(dataset: str, llm: str, model: str) -> Tuple[int, int]:
    """Return (done_count, total). Complete iff done >= total > 0."""
    d = model_out_dir(dataset, llm, model)
    total = dataset_size(dataset)
    # Both ablation wrappers delegate to SCOPE/run.py and produce predictions.jsonl.
    return _count_jsonl_lines(d / "predictions.jsonl"), total


def model_output_exists(dataset: str, llm: str, model: str) -> bool:
    done, total = model_output_progress(dataset, llm, model)
    return done >= total > 0


def missing_model_outputs(dataset: str, llm: str, models: List[str]) -> List[str]:
    return [m for m in models if not model_output_exists(dataset, llm, m)]


def available_eval_inputs(dataset: str, llm: str, models: List[str]) -> List[str]:
    base = out_root(dataset, llm)
    ready: List[str] = []
    for slug in models:
        p = predictions_path(base, dataset, slug)
        if file_has_content(p):
            ready.append(slug)
    return ready


def eval_summary_covers(summary_file: Path, dataset: str, llm: str, models: List[str]) -> bool:
    if not file_has_content(summary_file):
        return False
    ready = available_eval_inputs(dataset, llm, models)
    if not ready:
        return False
    try:
        with summary_file.open("r", encoding="utf-8") as f:
            done_keys = set((json.load(f).get("summaries") or {}).keys())
    except Exception:
        return False
    return all(MODEL_TO_SOURCE.get(slug, slug) in done_keys for slug in ready)


# ---------------------------------------------------------------------------
# Runners
# ---------------------------------------------------------------------------

def run_one_model(dataset: str, llm: str, model: str, force: bool, quiet: bool = False) -> Dict[str, Any]:
    out_dir = model_out_dir(dataset, llm, model)
    out_dir.mkdir(parents=True, exist_ok=True)

    done, total = model_output_progress(dataset, llm, model)
    if not force and done >= total > 0:
        banner(f"[skip] {dataset} / {llm} / {model}  ({done}/{total} done)")
        return {"status": "skipped", "elapsed": 0.0}

    if done > 0:
        banner(f"[resume] {dataset} / {llm} / {model}  ({done}/{total} done, continuing)")
    else:
        banner(f"[run]  {dataset} / {llm} / {model}  (0/{total})")
    cmd, cwd, env = MODEL_CMD_BUILDERS[model](dataset, llm, out_dir)
    rc, elapsed = run_cmd(cmd, log_path(dataset, llm, model), cwd=cwd, env=env, quiet=quiet)
    status = "ok" if rc == 0 else f"failed(rc={rc})"
    print(f"[done] {dataset}/{llm}/{model}: {status} in {elapsed:.0f}s", flush=True)
    return {"status": status, "elapsed": elapsed, "returncode": rc}


def run_eval(dataset: str, llm: str, force: bool) -> Dict[str, Any]:
    """Re-invoke the evaluator with ablation flags only.

    The full SCOPE (used as the reference row in the ablation table) is read
    later from the MAIN experiment's summary.json, so we do not pass --new-model
    here — only the ablation variants get judged against the gold set.
    """
    ds = DATASETS[dataset]
    eval_script = EVAL_DIR / ds["evaluator"]
    base = out_root(dataset, llm)
    eval_out_dir = base / "eval_summary"
    summary_file = eval_out_dir / "summary.json"

    if not force and eval_summary_covers(summary_file, dataset, llm, ABLATION_MODELS):
        banner(f"[skip-eval] {dataset} / {llm}  (summary.json up-to-date)")
        return {"status": "skipped", "summary": summary_file}

    input_slugs = available_eval_inputs(dataset, llm, ABLATION_MODELS)
    if not input_slugs:
        banner(f"[skip-eval] {dataset} / {llm}  (no ablation predictions found)")
        return {"status": "skipped-no-input", "summary": summary_file}

    banner(f"[eval] {dataset} / {llm}  (ablation)")
    eval_out_dir.mkdir(parents=True, exist_ok=True)

    cmd: List[str] = [
        "python3", str(eval_script),
        "--gold", str(ds["file"]),
        "--out-dir", str(eval_out_dir),
        "--llm-url", JUDGE_LLM_URL,
        "--llm-model", JUDGE_LLM_MODEL,
        "--api-key", JUDGE_API_KEY,
        "--max-workers", str(PER_MODEL_WORKERS),
    ]
    for slug in input_slugs:
        flag = MODEL_EVAL_FLAGS[slug]
        p = predictions_path(base, dataset, slug)
        cmd += [flag.flag, str(p)]

    rc, elapsed = run_cmd(cmd, eval_log_path(dataset, llm))
    status = "ok" if rc == 0 else f"failed(rc={rc})"
    print(f"[eval-done] {dataset}/{llm}: {status} in {elapsed:.0f}s", flush=True)
    return {"status": status, "elapsed": elapsed, "summary": summary_file}


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------

def _read_summary(path: Path) -> Optional[Dict[str, Any]]:
    if not file_has_content(path):
        return None
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"[warn] failed to read {path}: {e}", flush=True)
        return None


# All models shown in the ablation table — comes from eval/models_config.py.
ABLA_MODEL_ORDER = list(ABLATION_TABLE_MODELS)


def _summary_path_for(dataset: str, llm: str, model: str) -> Path:
    """SCOPE summary lives in the MAIN experiment dir; ablation summaries
    live in this script's ex_abla dir."""
    root = RESULT_ROOT_MAIN if model == "SCOPE" else RESULT_ROOT
    return root / dataset / _safe_name(llm) / "eval_summary" / "summary.json"


def _extract_cell(summary: Dict[str, Any], model: str) -> Tuple[Optional[float], Optional[float]]:
    """Returns (strict, loose) accuracy on stage sq2."""
    if not summary:
        return None, None
    src = MODEL_TO_SOURCE.get(model, model)
    entry = (summary.get("summaries") or {}).get(src)
    if not entry:
        return None, None
    sq2 = entry.get("sq2")
    if not sq2:
        return None, None
    return sq2.get("strict"), sq2.get("loose")


def build_tables(out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    overall: Dict[str, Any] = {}
    for dataset in DATASET_ORDER:
        rows: List[List[str]] = []
        header = ["model \\ LLM"] + LLM_ORDER
        rows.append(header)
        for model in ABLA_MODEL_ORDER:
            row = [model]
            for llm in LLM_ORDER:
                summary = _read_summary(_summary_path_for(dataset, llm, model))
                strict, loose = _extract_cell(summary or {}, model)
                if strict is None and loose is None:
                    row.append("-")
                else:
                    s = "-" if strict is None else f"{strict:.3f}"
                    l = "-" if loose  is None else f"{loose:.3f}"
                    row.append(f"{s} / {l}")
            rows.append(row)
        overall[dataset] = {"rows": rows}

    with (out_dir / "abla_tables.json").open("w", encoding="utf-8") as f:
        json.dump(overall, f, ensure_ascii=False, indent=2)

    build_excel_table(overall, out_dir)


def build_excel_table(all_tables: Dict[str, Any], out_dir: Path) -> None:
    """Write abla_tables.xlsx.

    Layout: one sheet per LLM that has any data.
    Each sheet:
      - Row 1 : merged title
      - Row 2 : column headers  (Model | KG-Doc | KG-Table | Table-Doc)
      - Row 3 : SCOPE (Full) — absolute scores, bold, gray background
      - Rows 4+: ablation variants — two-line cell:
                    line 1: strict / loose  (black, bold)
                    line 2: (Δstrict / Δloose) vs Full Model
                             red  if Δstrict < 0  (worse)
                             green if Δstrict > 0  (better)
                             gray  if no change
    """
    if not _OPENPYXL:
        print("[tables] openpyxl not installed; skipping Excel output (pip install openpyxl)")
        return

    DATASET_LABELS = {"kg_doc": "KG-Doc", "kg_table": "KG-Table", "table_doc": "Table-Doc"}

    def _val(ds: str, model: str, llm: str) -> str:
        rows = all_tables.get(ds, {}).get("rows", [])
        hdr = rows[0] if rows else []
        if llm not in hdr:
            return "-"
        col = hdr.index(llm)
        for row in rows[1:]:
            if row[0] == model:
                return row[col] if col < len(row) else "-"
        return "-"

    def _parse(v: str) -> Optional[Tuple[float, float]]:
        if v == "-":
            return None
        parts = v.split(" / ")
        if len(parts) != 2:
            return None
        try:
            return float(parts[0]), float(parts[1])
        except ValueError:
            return None

    def _thin():   return _XlSide(style="thin")
    def _medium(): return _XlSide(style="medium")
    def _brd(top=None, bottom=None, left=None, right=None):
        return _XlBorder(top=top, bottom=bottom, left=left, right=right)

    wb = _XlWorkbook()
    wb.remove(wb.active)

    created_sheet = False

    for llm in LLM_ORDER:
        # Skip LLMs with no data at all.
        if all(_val(ds, m, llm) == "-" for ds in DATASET_ORDER for m in ABLA_MODEL_ORDER):
            continue

        sheet_name = LLM_DISPLAY_NAMES.get(llm, llm)[:31]
        ws = wb.create_sheet(title=sheet_name)
        created_sheet = True

        ws.column_dimensions["A"].width = 28
        for col_letter in ["B", "C", "D"]:
            ws.column_dimensions[col_letter].width = 24

        # ── Row 1: title ──────────────────────────────────────────────────────
        ws.merge_cells("A1:D1")
        tc = ws["A1"]
        tc.value = (f"Ablation Results  ·  {LLM_DISPLAY_NAMES.get(llm, llm)}"
                    f"  (strict / loose accuracy,  Δ = ablation − full model)")
        tc.font      = _XlFont(name="Arial", bold=True, size=11)
        tc.alignment = _XlAlignment(horizontal="center", vertical="center")
        tc.border    = _brd(_medium(), _medium(), _medium(), _medium())
        ws.row_dimensions[1].height = 24

        # ── Row 2: column headers ─────────────────────────────────────────────
        headers = ["Model"] + [DATASET_LABELS[ds] for ds in DATASET_ORDER]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font      = _XlFont(name="Arial", bold=True, size=10)
            cell.alignment = _XlAlignment(horizontal="center", vertical="center")
            left_brd  = _medium() if ci == 1 else _thin()
            right_brd = _medium() if ci == 4 else _thin()
            cell.border = _brd(_medium(), _medium(), left_brd, right_brd)
        ws.row_dimensions[2].height = 18

        # ── Pre-compute full-model scores for delta calculation ───────────────
        full: Dict[str, Optional[Tuple[float, float]]] = {
            ds: _parse(_val(ds, "SCOPE", llm)) for ds in DATASET_ORDER
        }

        # ── Data rows ─────────────────────────────────────────────────────────
        n_models = len(ABLA_MODEL_ORDER)
        for row_i, model in enumerate(ABLA_MODEL_ORDER):
            cur      = 3 + row_i
            is_last  = (row_i == n_models - 1)
            is_full  = (model == "SCOPE")
            bot      = _medium() if is_last else _thin()
            ws.row_dimensions[cur].height = 16 if is_full else 30

            # Model name cell
            mc = ws.cell(row=cur, column=1,
                         value=ABLATION_DISPLAY_NAMES.get(model, model))
            mc.font      = _XlFont(name="Arial", size=10, bold=is_full)
            mc.alignment = _XlAlignment(horizontal="left", vertical="center",
                                        indent=1, wrap_text=False)
            mc.border    = _brd(_thin(), bot, _medium(), _thin())

            # Score cells (one per dataset)
            for col_idx, ds in enumerate(DATASET_ORDER, start=2):
                right_brd = _medium() if col_idx == 4 else _thin()
                cell = ws.cell(row=cur, column=col_idx)
                cell.border = _brd(_thin(), bot, _thin(), right_brd)

                v      = _val(ds, model, llm)
                parsed = _parse(v)

                if parsed is None:
                    # No data
                    cell.value     = "—"
                    cell.font      = _XlFont(name="Arial", size=10)
                    cell.alignment = _XlAlignment(horizontal="center", vertical="center")
                elif is_full:
                    # Full model: just the score, bold
                    cell.value     = f"{parsed[0]:.3f} / {parsed[1]:.3f}"
                    cell.font      = _XlFont(name="Arial", size=10, bold=True)
                    cell.alignment = _XlAlignment(horizontal="center", vertical="center")
                else:
                    # Ablation row: score + delta on second line
                    full_p = full.get(ds)
                    if full_p:
                        d_s = parsed[0] - full_p[0]
                        d_l = parsed[1] - full_p[1]
                        sign_s = "+" if d_s >= 0 else ""
                        sign_l = "+" if d_l >= 0 else ""
                        delta_str = f"({sign_s}{d_s:.3f} / {sign_l}{d_l:.3f})"
                        cell.value = f"{parsed[0]:.3f} / {parsed[1]:.3f}\n{delta_str}"
                        cell.font  = _XlFont(name="Arial", size=10)
                    else:
                        cell.value = f"{parsed[0]:.3f} / {parsed[1]:.3f}"
                        cell.font  = _XlFont(name="Arial", size=10)
                    cell.alignment = _XlAlignment(horizontal="center", vertical="center",
                                                  wrap_text=True)

        ws.freeze_panes = "B3"

    if not created_sheet:
        print("[tables] no ablation data available; skipping Excel output")
        return

    xlsx_path = out_dir / "abla_tables.xlsx"
    wb.save(xlsx_path)
    print(f"[tables] wrote {xlsx_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description="Ablation experiment: N LLMs × 3 datasets × multiple ablation variants")
    ap.add_argument("--only-llm",     choices=LLM_ORDER,       help="Restrict to this LLM")
    ap.add_argument("--only-dataset", choices=DATASET_ORDER,   help="Restrict to this dataset")
    ap.add_argument("--only-model",   choices=ABLATION_MODELS, help="Restrict to this ablation model")
    ap.add_argument("--skip-eval",   action="store_true", help="Don't run evaluators")
    ap.add_argument("--skip-runs",   action="store_true", help="Don't run any model — eval / tables only")
    ap.add_argument("--tables-only", action="store_true", help="Just aggregate existing summaries into tables")
    ap.add_argument("--force",       action="store_true", help="Re-run even if outputs exist")
    ap.add_argument("--serial-models", action="store_true",
                    help="Run ablation models serially within each (llm, dataset). "
                         "Default: run them in parallel.")
    ap.add_argument("--serial-datasets", action="store_true",
                    help="Process datasets one at a time.")
    args = ap.parse_args()

    RESULT_ROOT.mkdir(parents=True, exist_ok=True)
    LOG_ROOT.mkdir(parents=True, exist_ok=True)

    if args.tables_only:
        build_tables(RESULT_ROOT / "tables")
        return

    llms     = [args.only_llm]     if args.only_llm     else LLM_ORDER
    datasets = [args.only_dataset] if args.only_dataset else DATASET_ORDER
    models   = [args.only_model]   if args.only_model   else ABLATION_MODELS

    progress: Dict[str, Any] = {}
    progress_file = RESULT_ROOT / "_abla_progress.json"
    if progress_file.exists():
        try:
            progress = json.loads(progress_file.read_text(encoding="utf-8"))
        except Exception:
            progress = {}

    def save_progress():
        progress_file.write_text(
            json.dumps(progress, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _run_serial_dataset_loop(llm: str) -> None:
        for dataset in datasets:
            banner(f"### ABLA  LLM={llm}  DATASET={dataset}")
            progress.setdefault(llm, {}).setdefault(dataset, {})

            if not args.skip_runs:
                if args.serial_models or len(models) <= 1:
                    for model in models:
                        res = run_one_model(dataset, llm, model, force=args.force)
                        progress[llm][dataset][model] = res
                        save_progress()
                else:
                    banner(f"[parallel] launching {len(models)} ablation models concurrently "
                           f"for {dataset} / {llm}")
                    with ThreadPoolExecutor(max_workers=len(models)) as ex:
                        futs = {ex.submit(run_one_model, dataset, llm, m, args.force, quiet=True): m
                                for m in models}
                        for fut in as_completed(futs):
                            m = futs[fut]
                            try:
                                res = fut.result()
                            except Exception as e:
                                res = {"status": f"exception: {e!r}", "elapsed": 0.0}
                            progress[llm][dataset][m] = res
                            save_progress()

            if not args.skip_eval:
                missing = missing_model_outputs(dataset, llm, models)
                if missing and not args.skip_runs:
                    status = "skipped-missing-model-outputs"
                    print(f"[skip-eval] {dataset}/{llm}: missing complete outputs for {missing}", flush=True)
                    progress[llm][dataset]["__eval__"] = {"status": status, "missing": missing}
                    save_progress()
                else:
                    ev = run_eval(dataset, llm, force=args.force)
                    progress[llm][dataset]["__eval__"] = {
                        "status":  ev["status"], "elapsed": ev.get("elapsed"),
                        "summary": str(ev.get("summary")),
                    }
                    save_progress()

    def _run_parallel_dataset_loop(llm: str) -> None:
        """Default: launch every (dataset, model) for this LLM concurrently."""
        for dataset in datasets:
            progress.setdefault(llm, {}).setdefault(dataset, {})

        if args.skip_runs:
            if not args.skip_eval:
                with ThreadPoolExecutor(max_workers=len(datasets)) as eval_ex:
                    futs = {eval_ex.submit(run_eval, ds, llm, args.force): ds for ds in datasets}
                    for fut in as_completed(futs):
                        ds = futs[fut]
                        try:
                            ev = fut.result()
                        except Exception as e:
                            ev = {"status": f"exception: {e!r}"}
                        progress[llm][ds]["__eval__"] = {
                            "status": ev["status"], "elapsed": ev.get("elapsed"),
                            "summary": str(ev.get("summary")),
                        }
                        save_progress()
            return

        remaining = {ds: set(models) for ds in datasets}
        state_lock = threading.Lock()
        eval_futures: List[Tuple[str, Any]] = []
        n_model_slots = len(datasets) * len(models)

        banner(f"### ABLA  LLM={llm}  —  launching {len(datasets)} datasets × {len(models)} "
               f"ablation models  ({n_model_slots} concurrent; logs in {LOG_ROOT}/)")

        model_ex = ThreadPoolExecutor(max_workers=n_model_slots)
        eval_ex = ThreadPoolExecutor(max_workers=len(datasets))
        try:
            model_futs = {}
            for ds in datasets:
                for m in models:
                    f = model_ex.submit(run_one_model, ds, llm, m, args.force, quiet=True)
                    model_futs[f] = (ds, m)

            for fut in as_completed(model_futs):
                ds, m = model_futs[fut]
                try:
                    res = fut.result()
                except Exception as e:
                    res = {"status": f"exception: {e!r}", "elapsed": 0.0}
                fire_eval = False
                with state_lock:
                    progress[llm][ds][m] = res
                    save_progress()
                    remaining[ds].discard(m)
                    if not remaining[ds]:
                        fire_eval = True
                if fire_eval and not args.skip_eval:
                    missing = missing_model_outputs(ds, llm, models)
                    if missing:
                        print(f"[skip-eval] {ds}/{llm}: missing complete outputs for {missing}", flush=True)
                        progress[llm][ds]["__eval__"] = {
                            "status": "skipped-missing-model-outputs",
                            "missing": missing,
                        }
                        save_progress()
                    else:
                        print(f"[eval-trigger] {ds}/{llm}: all selected ablation outputs complete, "
                              f"launching evaluator", flush=True)
                        eval_futures.append((ds, eval_ex.submit(run_eval, ds, llm, args.force)))
        finally:
            model_ex.shutdown(wait=True)

        for ds, ef in eval_futures:
            try:
                ev = ef.result()
                progress[llm][ds]["__eval__"] = {
                    "status": ev["status"], "elapsed": ev.get("elapsed"),
                    "summary": str(ev.get("summary")),
                }
            except Exception as e:
                progress[llm][ds]["__eval__"] = {"status": f"exception: {e!r}"}
            save_progress()
        eval_ex.shutdown(wait=True)

    for llm in llms:
        if args.serial_datasets or args.serial_models:
            _run_serial_dataset_loop(llm)
        else:
            _run_parallel_dataset_loop(llm)

    build_tables(RESULT_ROOT / "tables")
    banner("ABLATION DONE")
    print(f"Tables: {RESULT_ROOT / 'tables'}")
    print(f"Progress log: {progress_file}")


if __name__ == "__main__":
    main()
