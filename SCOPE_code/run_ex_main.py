"""
Main experiment execution framework: N LLMs × 3 datasets × 11 models.

Models: new_model (ours) + 10 baselines
    Without retrieval: standard_prompt, cot, self_ask
    With retrieval:    standard_rag, ircot, cok, tog2, hydrarag, deepsieve, atomr

Layout (mirrors what the evaluators in /root/autodl-tmp/eval/ expect):
    /root/autodl-tmp/eval/result/
        <dataset_slug>/                 # kg_doc | kg_table | table_doc
            <llm_slug>/
                new_model/predictions.jsonl
                atomr/kgdoc_pred.jsonl  (+ trace/idx_*/final.trace.json for kg_table)
                deepsieve/query_*_results.jsonl  (+ traces/idx_*/...)
                hydrarag/predictions.jsonl
                eval_summary/summary.json
        tables/
            main_tables.xlsx

The runner is resumable: if an expected output file/dir is already present and
non-empty, the corresponding step is skipped. Use `--force` to re-run.

Recommended invocation (run yourself):

    cd /root/autodl-tmp && python3 run_ex_main.py
    cd /root/autodl-tmp && python3 run_ex_main.py --only-llm gpt-4o-mini
    cd /root/autodl-tmp && python3 run_ex_main.py --only-llm gpt-4o-mini --only-dataset kg_doc
    cd /root/autodl-tmp && python3 run_ex_main.py --tables-only
    cd /root/autodl-tmp && python3 run_ex_main.py --only-llm deepseek-chat --only-dataset kg_doc --only-model tog2 --skip-eval

By default the framework runs models *in parallel* within a (llm, dataset) combo.
Pass `--serial-models` to fall back to one-at-a-time.

cd /root/autodl-tmp
python3 run_ex_main.py \
  --only-llm deepseek-chat \
  --only-dataset kg_table \
  --serial-datasets \

  --serial-models \
  --skip-model tog2


"""
from __future__ import annotations

import argparse
import csv
import json
import os
import shlex
import subprocess
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

sys.path.insert(0, str(Path(__file__).resolve().parent / "eval"))
from models_config import (  # noqa: E402
    MAIN_EXPERIMENT_MODELS,
    MODEL_DISPLAY_NAMES,
    MODEL_EVAL_FLAGS,
    MODEL_TO_SOURCE,
    MAIN_WITHOUT_RETRIEVE,
    MAIN_WITH_RETRIEVE,
    MAIN_OUR_METHOD,
    predictions_path,
)

try:
    from openpyxl import Workbook as _XlWorkbook
    from openpyxl.styles import Font as _XlFont, Alignment as _XlAlignment
    from openpyxl.styles import Border as _XlBorder, Side as _XlSide
    from openpyxl.cell.rich_text import CellRichText as _XlRichText, TextBlock as _XlTextBlock
    from openpyxl.cell.text import InlineFont as _XlInlineFont
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


# ---------------------------------------------------------------------------
# Static configuration
# ---------------------------------------------------------------------------

QA_BENCH = Path("/root/autodl-tmp/new_model/qa_bench")
EVAL_DIR = Path("/root/autodl-tmp/eval")
RESULT_ROOT = EVAL_DIR / "result" / "ex_main"
LOG_ROOT = RESULT_ROOT / "_logs"
AOP_PYTHON = "/root/autodl-tmp/conda/envs/atomr/bin/python"

DATASETS: Dict[str, Dict[str, Any]] = {
    "kg_doc":    {"file": QA_BENCH / "kg-doc-1154.jsonl",  "kb": "kg,doc",    "evaluator": "evaluate_answer_kg_doc.py"},
    "kg_table":  {"file": QA_BENCH / "kg-table-1147.jsonl", "kb": "kg,table",  "evaluator": "evaluate_answer_kg_table.py"},
    "table_doc": {"file": QA_BENCH / "table-doc-1120.jsonl","kb": "table,doc", "evaluator": "evaluate_answer_table_doc.py"},
}

DATASET_ORDER = ["kg_doc", "kg_table", "table_doc"]
# Models that get run + evaluated for the main experiment.
# Edit eval/models_config.py to add/remove models.
MODEL_ORDER   = list(MAIN_EXPERIMENT_MODELS)
LLM_ORDER     = ["gpt-4o-mini", "gpt-4o", "qwen3-8b", "qwen3-32b", "qwen3-max-2026-01-23", "qwen3.5-397b-a17b", "gemini-2.5-flash", "glm-5", "deepseek-r1", "deepseek-chat"]

# The 7 new baselines share the same predictions.jsonl schema (sq1/sq2/final)
# as new_model. Three of them are pure-LLM (no retrieval), four go through
# the shared KG / Table-BM25 / Doc-ColBERT services with --kb.
SIMPLE_BASELINES_NO_RETRIEVAL = {"standard_prompt", "cot", "self_ask"}
SIMPLE_BASELINES_RETRIEVAL    = {"standard_rag", "ircot", "cok", "tog2"}
SIMPLE_BASELINES = SIMPLE_BASELINES_NO_RETRIEVAL | SIMPLE_BASELINES_RETRIEVAL

# Models that emit a cost_summary.json after their run. The 3 no-retrieval
# baselines (standard_prompt / cot / self_ask) are intentionally excluded —
# the user asked to skip cost stats for them.
COST_TRACKED_MODELS = {
    "new_model", "hydrarag", "atomr", "deepsieve", "aop",
    "standard_rag", "ircot", "cok", "tog2",
}
SIMPLE_BASELINE_SUBDIR = {
    "standard_prompt": "StandardPrompt",
    "cot":             "CoT",
    "self_ask":        "SelfAsk",
    "standard_rag":    "StandardRAG",
    "ircot":           "IRCoT",
    "cok":             "CoK",
    "tog2":            "ToG2",
}

# Main-experiment model groups and display names live in eval/models_config.py.
MAIN_DISPLAY_NAMES = MODEL_DISPLAY_NAMES

LLM_DISPLAY_NAMES: Dict[str, str] = {
    "gpt-4o-mini":                "GPT-4o-mini",
    "gpt-4o":                     "GPT-4o",
    "qwen3-8b":                   "Qwen3-8B",
    "qwen3-32b":                  "Qwen3-32B",
    "qwen3-max-2026-01-23":       "Qwen3-Max-2026-01-23",
    "qwen3.5-397b-a17b":          "Qwen3.5-397B-A17B",
    "gemini-2.5-flash":           "Gemini-2.5-Flash",
    "glm-5":                       "GLM-5",
    "deepseek-r1":                "DeepSeek-R1",
    "deepseek-chat":              "DeepSeek-Chat",
}

# Per-LLM × per-model API key configuration.
LLM_CONFIGS: Dict[str, Dict[str, Any]] = {
    "gpt-4o-mini": {
        "model": "gpt-4o-mini",
        "base_url": "https://.../v1",
        "keys": {
            "new_model": "sk-......",
            "atomr":     "sk-......",
            "ircot":     "sk-......",
            "deepsieve": "sk-......",
            "hydrarag":  "sk-......",
            "self_ask":  "sk-......"
        },
    },
    "gpt-4o": {
        "model": "gpt-4o",
        "base_url": "https://.../v1",
        "keys": {
            "new_model": "sk-......",
            "atomr":     "sk-......",
            "deepsieve": "sk-......",
            "hydrarag":  "sk-......",
        },
    },
    "qwen3-8b": {
        "model": "Qwen/Qwen3-8B",
        "base_url": "https://.../v1",
        "keys": {
            "new_model":       "sk-......",
            "atomr":           "sk-......",
            "deepsieve":       "sk-......",
            "hydrarag":        "sk-......",
            "standard_prompt": "sk-......",
            "cot":             "sk-......",
            "self_ask":        "sk-......",
            "standard_rag":    "sk-......",
            "ircot":           "sk-......",
            "cok":             "sk-......",
            "tog2":            "sk-......",
        },
    },
    "qwen3-32b": {
        "model": "Qwen3-32B",
        "base_url": "https://.../v1",
        "keys": {
            "new_model": "sk-......",
            "atomr":     "sk-......",
            "deepsieve": "sk-......",
            "hydrarag":  "sk-......",
            "ircot":     "sk-......",
            "self_ask":  "sk-......",
        },
    },
    "qwen3-max-2026-01-23": {
        "model": "qwen3-max-2026-01-23",
        "base_url": "https://.../v1",
        "keys": {
            "new_model": "sk-......",
            "atomr":     "sk-......",
            "ircot":     "sk-......",
            "deepsieve": "sk-......",
            "hydrarag":  "sk-......",
            "self_ask":  "sk-......"
        },
    },
    "qwen3.5-397b-a17b": {
        "model": "qwen3.5-397b-a17b",
        "base_url": "https://.../v1",
        "keys": {
            "new_model": "sk-......",
            "atomr":     "sk-......",
            "ircot":     "sk-......",
            "deepsieve": "sk-......",
            "hydrarag":  "sk-......",
            "self_ask":  "sk-......"
        },
    },
    "gemini-2.5-flash": {
        "model": "gemini-2.5-flash",
        "base_url": "https://.../v1",
        "keys": {
            "new_model": "sk-......",
            "atomr":     "sk-......",
            "ircot":     "sk-......",
            "deepsieve": "sk-......",
            "hydrarag":  "sk-......",
            "self_ask":  "sk-......"
        },
    },
    "glm-5": {
        "model": "GLM-5",
        "base_url": "https://.../v1",
        "keys": {
            "new_model": "sk-......",
            "atomr":     "sk-......",
            "deepsieve": "sk-......",
            "hydrarag":  "sk-......",
            "ircot":     "sk-......",
            "self_ask":  "sk-......",
        },
    },
    "deepseek-r1": {
        "model": "DeepSeek-R1-0528",
        "base_url": "https://.../v1",
        "keys": {
            "new_model": "sk-......",
            "atomr":     "sk-......",
            "deepsieve": "sk-......",
            "hydrarag":  "sk-......",
            "ircot":     "sk-......",
            "self_ask":  "sk-......",
        },
    },
    "deepseek-chat": {
        "model": "deepseek-chat",
        "base_url": "https://.../v1",
        # 3 DeepSeek keys evenly distributed across the 11 main models.
        "keys": {
            # ---- KEY_A: sk-bafb98... ----
            "new_model":       "sk-......",
            "hydrarag":        "sk-......",
            "ircot":           "sk-......",
            "standard_prompt": "sk-......",
            "cot":             "sk-......",
            # ---- KEY_B: sk-166859... ----
            "atomr":     "sk-......",
            "cok":       "sk-......",
            "self_ask":  "sk-......",
            # ---- KEY_C: sk-b0fa7b... ----
            "deepsieve":    "sk-......",
            "tog2":         "sk-......",
            "standard_rag": "sk-......",
        },
    },
}

# Fill in keys for the 7 simple baselines + AOP on every LLM config that doesn't
# already list them — fall back to the new_model key.
for _cfg in LLM_CONFIGS.values():
    _keys = _cfg["keys"]
    _fallback = _keys.get("new_model") or next(iter(_keys.values()))
    for _m in ("standard_prompt", "cot", "self_ask",
               "standard_rag", "ircot", "cok", "tog2",
               "aop"):
        _keys.setdefault(_m, _fallback)


# Evaluator (LLM-judge) — fixed across all runs.
# JUDGE_LLM_MODEL = "DeepSeek-V4-Flash"
# JUDGE_LLM_URL   = "https://www.sophnet.com/api/open-apis/v1"
# JUDGE_API_KEY   = "V8IjyYmmJLK4vniImZ9IpJaowdnjAIR1s84Ch8sDWQQLpqm7TJaGyp2atttRh7hXg54l2H6GYzpBHZQDEQC2wQ"
JUDGE_LLM_MODEL = "deepseek-chat"
JUDGE_LLM_URL   = "https://.../v1"
JUDGE_API_KEYS_BY_DATASET = {
    "kg_doc":   "sk-......",
    "kg_table": "sk-......",
}
JUDGE_MAX_WORKERS = 24

# Concurrency hint for each model's own --workers / --concurrency flag.
PER_MODEL_WORKERS = 3
PER_LLM_WORKERS: Dict[str, int] = {
    "deepseek-chat": 24,
    "gpt-4o-mini": 20,
    "qwen3-8b": 3,
    "qwen3-32b": 3,
    "qwen3-max-2026-01-23": 20,
    "qwen3.5-397b-a17b": 10,
    "gemini-2.5-flash": 20,
    "glm-5": 3,
    "deepseek-r1": 3,
}


def workers_for(llm: str) -> int:
    return PER_LLM_WORKERS.get(llm, PER_MODEL_WORKERS)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_name(s: str) -> str:
    return s.replace("/", "_").replace(" ", "_")


def out_root(dataset: str, llm: str) -> Path:
    return RESULT_ROOT / dataset / _safe_name(llm)


def model_out_dir(dataset: str, llm: str, model: str) -> Path:
    return out_root(dataset, llm) / model


def log_path(dataset: str, llm: str, model: str) -> Path:
    LOG_ROOT.mkdir(parents=True, exist_ok=True)
    return LOG_ROOT / f"{dataset}__{_safe_name(llm)}__{model}.log"


def eval_log_path(dataset: str, llm: str) -> Path:
    LOG_ROOT.mkdir(parents=True, exist_ok=True)
    return LOG_ROOT / f"{dataset}__{_safe_name(llm)}__eval.log"


def banner(msg: str) -> None:
    bar = "=" * 78
    print(f"\n{bar}\n{msg}\n{bar}", flush=True)


def file_has_content(p: Path) -> bool:
    return p.exists() and p.is_file() and p.stat().st_size > 0


def dir_has_files(p: Path, pattern: str = "*") -> bool:
    return p.exists() and p.is_dir() and any(p.glob(pattern))


def run_cmd(
    cmd: List[str],
    log_file: Path,
    cwd: Optional[Path] = None,
    env: Optional[Dict[str, str]] = None,
    timeout: Optional[int] = None,
    quiet: bool = False,
) -> Tuple[int, float]:
    """Run a subprocess, tee output to a log file, return (returncode, elapsed).

    When `quiet=True`, child stdout/stderr is written only to the log file (no
    forwarding to the parent stdout). This avoids interleaved garbage when
    multiple subprocesses run concurrently.
    """
    log_file.parent.mkdir(parents=True, exist_ok=True)
    print(f"[exec] {' '.join(shlex.quote(c) for c in cmd)}", flush=True)
    print(f"[exec] cwd={cwd}  log={log_file}", flush=True)
    t0 = time.time()
    with log_file.open("w", encoding="utf-8") as lf:
        lf.write(f"# CMD: {' '.join(shlex.quote(c) for c in cmd)}\n")
        if cwd:
            lf.write(f"# CWD: {cwd}\n")
        lf.flush()
        try:
            proc = subprocess.Popen(
                cmd, cwd=str(cwd) if cwd else None,
                env={**os.environ, **(env or {})},
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                bufsize=1, universal_newlines=True,
                encoding="utf-8", errors="replace",
            )
        except FileNotFoundError as e:
            lf.write(f"\n[ERROR] {e}\n")
            return 127, time.time() - t0

        assert proc.stdout is not None
        for line in proc.stdout:
            if not quiet:
                sys.stdout.write(line)
                sys.stdout.flush()
            lf.write(line)
        proc.wait(timeout=timeout)
    elapsed = time.time() - t0
    return proc.returncode or 0, elapsed


# ---------------------------------------------------------------------------
# Per-model command builders
# ---------------------------------------------------------------------------
def cmd_new_model(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        "python3", "run.py",
        "--input", str(ds["file"]),
        "--gold",  str(ds["file"]),
        "--output-dir", str(out_dir),
        "--kb", ds["kb"],
        "--routing-mode", "graph",
        "--workers", str(workers_for(llm)),  # 改为这一行
        "--api-key",  cfg["keys"]["new_model"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
    ]
    return cmd, Path("/root/autodl-tmp/new_model"), {}

def cmd_hydrarag(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        "python3", "run.py",
        "--input", str(ds["file"]),
        "--output-dir", str(out_dir),
        "--kb", ds["kb"],
        "--workers", str(workers_for(llm)),
        "--api-key",  cfg["keys"]["hydrarag"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
    ]
    return cmd, Path("/root/autodl-tmp/baseline/HydraRAG/hydra_baseline"), {}


def cmd_atomr(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    trace_dir = out_dir / "trace"
    cmd = [
        "python3", "main_mline.py",
        "--dataset-name", "mmqa",
        "--dataset-path", str(ds["file"]),
        "--output-trees-path",       str(out_dir / f"{dataset}_trees.jsonl"),
        "--output-predictions-path", str(out_dir / f"{dataset}_pred.jsonl"),
        "--kb", ds["kb"],
        "--workers", str(workers_for(llm)),
        "--trace-dir", str(trace_dir),
        "--resume",
    ]
    env = {
        "OPENAI_API_KEY":   cfg["keys"]["atomr"],
        "OPENAI_BASE_URL":  cfg["base_url"],
        "ATOMR_LLM_MODEL":  cfg["model"],
        "ATOMR_KG_PARSER_MODEL": cfg["model"],
        "ATOMR_TRACE_DIR":  str(trace_dir),
    }
    return cmd, Path("/root/autodl-tmp/baseline/Atomr/src"), env


def cmd_aop(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        AOP_PYTHON, "run.py",
        "--input", str(ds["file"]),
        "--output-dir", str(out_dir),
        "--kb", ds["kb"],
        "--workers", str(workers_for(llm)),
        "--api-key",  cfg["keys"]["aop"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
        "--resume",
    ]
    env = {
        "HF_ENDPOINT": "https://hf-mirror.com",
    }
    return cmd, Path("/root/autodl-tmp/baseline/AOP/main"), env


def cmd_deepsieve(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        "python3", "runner/main_rag_only.py",
        "--dataset", "mmqa",
        "--dataset-path", str(ds["file"]),
        "--kb", ds["kb"],
        "--output-dir", str(out_dir),
        "--use_routing", "--decompose", "--use_reflection",
        "--max_reflexion_times", "2",
        "--concurrency", str(workers_for(llm)),
        "--sample_size", str(dataset_size(dataset)),
        "--api-key",  cfg["keys"]["deepsieve"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
        "--resume",
    ]
    return cmd, Path("/root/autodl-tmp/baseline/DeepSieve"), {}


def _make_simple_baseline_cmd(method_slug: str):
    """Factory: returns a cmd builder for one of the 7 simple baselines."""
    subdir = SIMPLE_BASELINE_SUBDIR[method_slug]
    needs_retrieval = method_slug in SIMPLE_BASELINES_RETRIEVAL

    def builder(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
        cfg = LLM_CONFIGS[llm]
        ds = DATASETS[dataset]
        cmd = [
            "python3", "run.py",
            "--input", str(ds["file"]),
            "--output-dir", str(out_dir),
            "--workers", str(workers_for(llm)),
            "--api-key",  cfg["keys"][method_slug],
            "--llm-url",  cfg["base_url"],
            "--llm-model", cfg["model"],
            "--resume",
        ]
        if needs_retrieval:
            cmd += ["--kb", ds["kb"]]
        cwd = Path(f"/root/autodl-tmp/baseline/{subdir}")
        return cmd, cwd, {}
    return builder


MODEL_CMD_BUILDERS = {
    "new_model": cmd_new_model,
    "hydrarag":  cmd_hydrarag,
    "atomr":     cmd_atomr,
    "deepsieve": cmd_deepsieve,
    "aop":       cmd_aop,
    **{m: _make_simple_baseline_cmd(m) for m in SIMPLE_BASELINES},
}


# ---------------------------------------------------------------------------
# Completion sentinels (used to decide "is this run already done?")
# ---------------------------------------------------------------------------

@lru_cache(maxsize=None)
def dataset_size(dataset: str) -> int:
    gold = DATASETS[dataset]["file"]
    n = 0
    with gold.open("rb") as f:
        for _ in f:
            n += 1
    return n


def _count_jsonl_lines(p: Path) -> int:
    if not p.exists() or not p.is_file():
        return 0
    n = 0
    with p.open("rb") as f:
        for _ in f:
            n += 1
    return n


def model_output_progress(dataset: str, llm: str, model: str) -> Tuple[int, int]:
    """Return (done_count, total). Complete iff done >= total > 0."""
    d = model_out_dir(dataset, llm, model)
    total = dataset_size(dataset)
    if model == "new_model":
        return _count_jsonl_lines(d / "predictions.jsonl"), total
    if model == "hydrarag":
        return _count_jsonl_lines(d / "predictions.jsonl"), total
    if model == "atomr":
        pred = _count_jsonl_lines(d / f"{dataset}_pred.jsonl")
        if dataset == "kg_table":
            trace = d / "trace"
            traces = len(list(trace.glob("idx_*"))) if trace.exists() else 0
            return max(pred, traces), total
        return pred, total
    if model == "deepsieve":
        files = list(d.glob("query_*_results.jsonl")) if d.exists() else []
        return len(files), total
    if model == "aop":
        return _count_jsonl_lines(d / "predictions.jsonl"), total
    if model in SIMPLE_BASELINES:
        return _count_jsonl_lines(d / "predictions.jsonl"), total
    return 0, total


def model_output_exists(dataset: str, llm: str, model: str) -> bool:
    done, total = model_output_progress(dataset, llm, model)
    return done >= total > 0


# ---------------------------------------------------------------------------
# Runners
# ---------------------------------------------------------------------------

def _read_cost_summary(out_dir: Path) -> Optional[Dict[str, Any]]:
    """Look for a cost_summary.json the model wrote — return parsed totals or None."""
    p = out_dir / "cost_summary.json"
    if not p.exists() or p.stat().st_size == 0:
        return None
    try:
        with p.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"[cost] failed to read {p}: {e}", flush=True)
        return None


def _format_cost_line(cost: Dict[str, Any]) -> str:
    llm = cost.get("llm_calls", 0)
    ret = cost.get("retrieval", {}) or {}
    tot_ret = cost.get("total_retrieval", sum(ret.values()))
    ret_str = " ".join(f"{k}:{v}" for k, v in sorted(ret.items())) or "(none)"
    return f"LLM calls = {llm}  |  retrieval = {ret_str}  (total {tot_ret})"


def _print_model_cost(dataset: str, llm: str, model: str, out_dir: Path) -> Dict[str, Any]:
    """Locate the model's cost_summary.json and banner-print it. Skipped for
    no-retrieval baselines (per user). Returns the cost dict (or empty)."""
    tag = f"{dataset}/{llm}/{model}"
    if model not in COST_TRACKED_MODELS:
        print(f"[cost] {tag}: skipped (no-retrieval baseline)", flush=True)
        return {}
    cost = _read_cost_summary(out_dir)
    if cost is None:
        print(f"[cost] {tag}: cost_summary.json missing or empty in {out_dir}", flush=True)
        return {}
    print(f"[cost] {tag}: {_format_cost_line(cost)}", flush=True)
    return cost


def run_one_model(dataset: str, llm: str, model: str, force: bool, quiet: bool = False) -> Dict[str, Any]:
    out_dir = model_out_dir(dataset, llm, model)
    out_dir.mkdir(parents=True, exist_ok=True)

    done, total = model_output_progress(dataset, llm, model)
    if not force and done >= total > 0:
        banner(f"[skip] {dataset} / {llm} / {model}  ({done}/{total} done)")
        cost = _print_model_cost(dataset, llm, model, out_dir)
        return {"status": "skipped", "elapsed": 0.0, "cost": cost}

    if done > 0:
        banner(f"[resume] {dataset} / {llm} / {model}  ({done}/{total} done, continuing)")
    else:
        banner(f"[run]  {dataset} / {llm} / {model}  (0/{total})")
    cmd, cwd, env = MODEL_CMD_BUILDERS[model](dataset, llm, out_dir)
    rc, elapsed = run_cmd(cmd, log_path(dataset, llm, model), cwd=cwd, env=env, quiet=quiet)
    status = "ok" if rc == 0 else f"failed(rc={rc})"
    print(f"[done] {dataset}/{llm}/{model}: {status} in {elapsed:.0f}s", flush=True)
    cost = _print_model_cost(dataset, llm, model, out_dir)
    return {"status": status, "elapsed": elapsed, "returncode": rc, "cost": cost}


def _eval_is_up_to_date(base: Path, dataset: str, summary_file: Path) -> bool:
    """Return True only if summary.json exists and already covers every model that has predictions."""
    if not file_has_content(summary_file):
        return False
    try:
        with summary_file.open() as f:
            done_keys = set((json.load(f).get("summaries") or {}).keys())
    except Exception:
        return False

    fixed_checks = [
        (file_has_content(base / "new_model" / "predictions.jsonl"),        "new_model"),
        (file_has_content(base / "hydrarag"  / "predictions.jsonl"),        "hyprarag"),
        (file_has_content(base / "atomr"     / f"{dataset}_pred.jsonl"),    "atomr"),
        (bool(list((base / "deepsieve").glob("query_*_results.jsonl")))
         if (base / "deepsieve").exists() else False,                        "deepserive"),
    ]
    for has_pred, key in fixed_checks:
        if has_pred and key not in done_keys:
            return False

    for slug in SIMPLE_BASELINES:
        p = base / slug / "predictions.jsonl"
        if file_has_content(p) and slug not in done_keys:
            return False

    return True


def run_eval(dataset: str, llm: str, force: bool) -> Dict[str, Any]:
    """Invoke the dataset-specific evaluator over all model outputs."""
    ds = DATASETS[dataset]
    eval_script = EVAL_DIR / ds["evaluator"]
    base = out_root(dataset, llm)
    eval_out_dir = base / "eval_summary"
    summary_file = eval_out_dir / "summary.json"

    if not force and _eval_is_up_to_date(base, dataset, summary_file):
        banner(f"[skip-eval] {dataset} / {llm}  (summary.json up-to-date)")
        return {"status": "skipped", "summary": summary_file}

    banner(f"[eval] {dataset} / {llm}")
    eval_out_dir.mkdir(parents=True, exist_ok=True)

    cmd: List[str] = [
        "python3", str(eval_script),
        "--gold", str(ds["file"]),
        "--out-dir", str(eval_out_dir),
        "--llm-url", JUDGE_LLM_URL,
        "--llm-model", JUDGE_LLM_MODEL,
        "--api-key", JUDGE_API_KEYS_BY_DATASET.get(dataset, JUDGE_API_KEYS_BY_DATASET["kg_doc"]),
        "--max-workers", str(JUDGE_MAX_WORKERS),
    ]
    # Pass --<flag> <path> for every main-experiment model whose predictions
    # are present. Slugs / flags / file patterns live in eval/models_config.py.
    for slug in MAIN_EXPERIMENT_MODELS:
        flag = MODEL_EVAL_FLAGS[slug]
        p = predictions_path(base, dataset, slug)
        if flag.is_dir:
            if p.exists() and any(p.glob("query_*_results.jsonl")):
                cmd += [flag.flag, str(p)]
        elif file_has_content(p):
            cmd += [flag.flag, str(p)]
    rc, elapsed = run_cmd(cmd, eval_log_path(dataset, llm))
    status = "ok" if rc == 0 else f"failed(rc={rc})"
    print(f"[eval-done] {dataset}/{llm}: {status} in {elapsed:.0f}s", flush=True)
    return {"status": status, "elapsed": elapsed, "summary": summary_file}


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------

def _read_summary(path: Path) -> Optional[Dict[str, Any]]:
    if not file_has_content(path):
        return None
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"[warn] failed to read {path}: {e}", flush=True)
        return None


# MODEL_TO_SOURCE and per-model eval flags are imported from eval/models_config.py.


def _extract_cell(summary: Dict[str, Any], model: str) -> Tuple[Optional[float], Optional[float]]:
    """Returns (strict, loose) accuracy on stage sq2 (the final-answer stage)."""
    if not summary:
        return None, None
    src = MODEL_TO_SOURCE[model]
    entry = (summary.get("summaries") or {}).get(src)
    if not entry:
        return None, None
    sq2 = entry.get("sq2")
    if not sq2:
        return None, None
    return sq2.get("strict"), sq2.get("loose")


def build_tables(out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    overall: Dict[str, Any] = {}
    for dataset in DATASET_ORDER:
        rows: List[List[str]] = []
        header = ["model \\ LLM"] + LLM_ORDER
        rows.append(header)
        for model in MODEL_ORDER:
            row = [model]
            for llm in LLM_ORDER:
                summary = _read_summary(out_root(dataset, llm) / "eval_summary" / "summary.json")
                strict, loose = _extract_cell(summary or {}, model)
                if strict is None and loose is None:
                    row.append("-")
                else:
                    s = "-" if strict is None else f"{strict:.3f}"
                    l = "-" if loose  is None else f"{loose:.3f}"
                    row.append(f"{s} / {l}")
            rows.append(row)

        overall[dataset] = {"rows": rows}

    with (out_dir / "all_tables.json").open("w", encoding="utf-8") as f:
        json.dump(overall, f, ensure_ascii=False, indent=2)

    build_excel_table(overall, out_dir)


def build_excel_table(all_tables: Dict[str, Any], out_dir: Path) -> None:
    """Write main_tables.xlsx — one sheet per LLM, rows=models (3 groups), cols=datasets."""
    if not _OPENPYXL:
        print("[tables] openpyxl not installed; skipping Excel output (pip install openpyxl)")
        return

    dataset_headers = ["KG-Doc", "KG-Table", "Table-Doc"]

    def _val(ds: str, model: str, llm: str) -> str:
        rows = all_tables.get(ds, {}).get("rows", [])
        hdr = rows[0] if rows else []
        if llm not in hdr:
            return "-"
        col = hdr.index(llm)
        for row in rows[1:]:
            if row[0] == model:
                return row[col] if col < len(row) else "-"
        return "-"

    def _parse(v: str) -> Optional[Tuple[float, float]]:
        if v == "-":
            return None
        parts = v.split(" / ")
        if len(parts) != 2:
            return None
        try:
            return float(parts[0]), float(parts[1])
        except ValueError:
            return None

    groups = [
        ("Without\nRetrieve", MAIN_WITHOUT_RETRIEVE),
        ("With\nRetrieve",    MAIN_WITH_RETRIEVE),
        ("Ours",              MAIN_OUR_METHOD),
    ]
    all_models = MAIN_WITHOUT_RETRIEVE + MAIN_WITH_RETRIEVE + MAIN_OUR_METHOD

    def _thin():   return _XlSide(style="thin")
    def _medium(): return _XlSide(style="medium")

    def _brd(top=None, bottom=None, left=None, right=None):
        return _XlBorder(top=top, bottom=bottom, left=left, right=right)

    wb = _XlWorkbook()
    wb.remove(wb.active)

    for llm in LLM_ORDER:
        if all(_val(ds, m, llm) == "-" for ds in DATASET_ORDER for m in all_models):
            continue

        max_strict: Dict[str, Optional[float]] = {}
        max_loose:  Dict[str, Optional[float]] = {}
        for ds in DATASET_ORDER:
            sv, lv = [], []
            for m in all_models:
                p = _parse(_val(ds, m, llm))
                if p:
                    sv.append(p[0]); lv.append(p[1])
            max_strict[ds] = max(sv) if sv else None
            max_loose[ds]  = max(lv) if lv else None

        sheet_name = LLM_DISPLAY_NAMES.get(llm, llm)[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 22
        for col in ["C", "D", "E"]:
            ws.column_dimensions[col].width = 18

        ws.merge_cells("A1:E1")
        tc = ws["A1"]
        tc.value = f"Results with {LLM_DISPLAY_NAMES.get(llm, llm)}  (strict / loose accuracy)"
        tc.font = _XlFont(name="Arial", bold=True, size=12)
        tc.alignment = _XlAlignment(horizontal="center", vertical="center")
        tc.border = _brd(_medium(), _medium(), _medium(), _medium())
        ws.row_dimensions[1].height = 24

        for ci, h in enumerate(["Category", "Model"] + dataset_headers, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font = _XlFont(name="Arial", bold=True, size=10)
            c.alignment = _XlAlignment(horizontal="center", vertical="center")
            c.border = _brd(_thin(), _medium(), _thin(), _thin())
        ws.row_dimensions[2].height = 18

        cur = 3
        for _g_idx, (cat_label, group) in enumerate(groups):
            n = len(group)
            group_start = cur
            for i, model in enumerate(group):
                is_last = i == n - 1
                bot = _medium() if is_last else _thin()
                ws.row_dimensions[cur].height = 16

                cc = ws.cell(row=cur, column=1, value=cat_label if i == 0 else None)
                cc.font = _XlFont(name="Arial", size=10)
                cc.alignment = _XlAlignment(horizontal="center", vertical="center", wrap_text=True)
                cc.border = _brd(
                    top=_thin() if i == 0 else None,
                    bottom=_medium() if is_last else None,
                    left=_thin(), right=_thin(),
                )

                mc = ws.cell(row=cur, column=2, value=MAIN_DISPLAY_NAMES[model])
                mc.font = _XlFont(name="Arial", size=10)
                mc.alignment = _XlAlignment(horizontal="center", vertical="center")
                mc.border = _brd(_thin(), bot, _thin(), _thin())

                for col_idx, ds in enumerate(DATASET_ORDER, start=3):
                    v = _val(ds, model, llm)
                    c = ws.cell(row=cur, column=col_idx)
                    c.alignment = _XlAlignment(horizontal="center", vertical="center")
                    c.border = _brd(_thin(), bot, _thin(), _thin())
                    parsed = _parse(v)
                    if parsed is None:
                        c.value = v
                        c.font = _XlFont(name="Arial", size=10)
                    else:
                        sv_f, lv_f = parsed
                        s_bold = max_strict[ds] is not None and abs(sv_f - max_strict[ds]) < 1e-9
                        l_bold = max_loose[ds]  is not None and abs(lv_f - max_loose[ds])  < 1e-9
                        s_str, l_str = f"{sv_f:.3f}", f"{lv_f:.3f}"
                        if not s_bold and not l_bold:
                            c.value = f"{s_str} / {l_str}"
                            c.font = _XlFont(name="Arial", size=10)
                        else:
                            c.value = _XlRichText(
                                _XlTextBlock(_XlInlineFont(b=s_bold, sz=10), s_str),
                                " / ",
                                _XlTextBlock(_XlInlineFont(b=l_bold, sz=10), l_str),
                            )
                cur += 1

            if n > 1:
                ws.merge_cells(start_row=group_start, start_column=1,
                               end_row=cur - 1, end_column=1)

        ws.freeze_panes = "A3"

    xlsx_path = out_dir / "main_tables.xlsx"
    wb.save(xlsx_path)
    print(f"[tables] wrote {xlsx_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description="Main experiment: N LLMs × 3 datasets × 11 models")
    ap.add_argument("--only-llm",     choices=LLM_ORDER,     help="Restrict to this LLM")
    ap.add_argument("--only-dataset", choices=DATASET_ORDER, help="Restrict to this dataset")
    ap.add_argument("--only-model",   choices=MODEL_ORDER,   help="Restrict to this model")
    ap.add_argument("--skip-model",   action="append", choices=MODEL_ORDER, default=[],
                    help="Skip one or more models. Repeat the flag to skip multiple models.")
    ap.add_argument("--skip-eval",  action="store_true", help="Don't run evaluators")
    ap.add_argument("--skip-runs",  action="store_true", help="Don't run any model — eval / tables only")
    ap.add_argument("--tables-only", action="store_true", help="Just aggregate existing summaries into tables")
    ap.add_argument("--force",       action="store_true", help="Re-run even if outputs exist")
    ap.add_argument("--serial-models", action="store_true",
                    help="Run models serially within each (llm, dataset). "
                         "Default: run them in parallel (one thread per model).")
    ap.add_argument("--serial-datasets", action="store_true",
                    help="Process datasets one at a time (legacy behavior). "
                         "Default: launch all (dataset × model) combos for a given LLM "
                         "concurrently and fire each dataset's evaluator as soon as its "
                         "models finish.")
    args = ap.parse_args()

    RESULT_ROOT.mkdir(parents=True, exist_ok=True)
    LOG_ROOT.mkdir(parents=True, exist_ok=True)

    if args.tables_only:
        build_tables(RESULT_ROOT / "tables")
        return

    llms     = [args.only_llm]     if args.only_llm     else LLM_ORDER
    datasets = [args.only_dataset] if args.only_dataset else DATASET_ORDER
    if args.only_model:
        models = [args.only_model]
    else:
        skipped_models = set(args.skip_model)
        models = [m for m in MODEL_ORDER if m not in skipped_models]

    progress: Dict[str, Any] = {}
    progress_file = RESULT_ROOT / "_progress.json"
    if progress_file.exists():
        try:
            progress = json.loads(progress_file.read_text(encoding="utf-8"))
        except Exception:
            progress = {}

    def save_progress():
        progress_file.write_text(
            json.dumps(progress, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _run_serial_dataset_loop(llm: str) -> None:
        for dataset in datasets:
            banner(f"### LLM={llm}  DATASET={dataset}")
            progress.setdefault(llm, {}).setdefault(dataset, {})

            if not args.skip_runs:
                if args.serial_models or len(models) <= 1:
                    for model in models:
                        res = run_one_model(dataset, llm, model, force=args.force)
                        progress[llm][dataset][model] = res
                        save_progress()
                else:
                    banner(f"[parallel] launching {len(models)} models concurrently "
                           f"for {dataset} / {llm}")
                    with ThreadPoolExecutor(max_workers=len(models)) as ex:
                        futs = {ex.submit(run_one_model, dataset, llm, m, args.force, quiet=True): m
                                for m in models}
                        for fut in as_completed(futs):
                            m = futs[fut]
                            try:
                                res = fut.result()
                            except Exception as e:
                                res = {"status": f"exception: {e!r}", "elapsed": 0.0}
                            progress[llm][dataset][m] = res
                            save_progress()

            if not args.skip_eval:
                ev = run_eval(dataset, llm, force=args.force)
                progress[llm][dataset]["__eval__"] = {
                    "status":  ev["status"], "elapsed": ev.get("elapsed"),
                    "summary": str(ev.get("summary")),
                }
                save_progress()

    def _run_parallel_dataset_loop(llm: str) -> None:
        """Default: launch every (dataset, model) for this LLM concurrently."""
        for dataset in datasets:
            progress.setdefault(llm, {}).setdefault(dataset, {})

        if args.skip_runs:
            if not args.skip_eval:
                with ThreadPoolExecutor(max_workers=len(datasets)) as eval_ex:
                    futs = {eval_ex.submit(run_eval, ds, llm, args.force): ds for ds in datasets}
                    for fut in as_completed(futs):
                        ds = futs[fut]
                        try:
                            ev = fut.result()
                        except Exception as e:
                            ev = {"status": f"exception: {e!r}"}
                        progress[llm][ds]["__eval__"] = {
                            "status": ev["status"], "elapsed": ev.get("elapsed"),
                            "summary": str(ev.get("summary")),
                        }
                        save_progress()
            return

        remaining = {ds: set(models) for ds in datasets}
        state_lock = threading.Lock()
        eval_futures: List[Tuple[str, Any]] = []
        n_model_slots = len(datasets) * len(models)

        banner(f"### LLM={llm}  —  launching {len(datasets)} datasets × {len(models)} "
               f"models  ({n_model_slots} concurrent subprocesses; per-model log in "
               f"{LOG_ROOT}/)")

        model_ex = ThreadPoolExecutor(max_workers=n_model_slots)
        eval_ex = ThreadPoolExecutor(max_workers=len(datasets))
        try:
            model_futs = {}
            for ds in datasets:
                for m in models:
                    f = model_ex.submit(run_one_model, ds, llm, m, args.force, quiet=True)
                    model_futs[f] = (ds, m)

            for fut in as_completed(model_futs):
                ds, m = model_futs[fut]
                try:
                    res = fut.result()
                except Exception as e:
                    res = {"status": f"exception: {e!r}", "elapsed": 0.0}
                fire_eval = False
                with state_lock:
                    progress[llm][ds][m] = res
                    save_progress()
                    remaining[ds].discard(m)
                    if not remaining[ds]:
                        fire_eval = True
                if fire_eval and not args.skip_eval:
                    print(f"[eval-trigger] {ds}/{llm}: all {len(models)} models done, "
                          f"launching evaluator", flush=True)
                    eval_futures.append((ds, eval_ex.submit(run_eval, ds, llm, args.force)))
        finally:
            model_ex.shutdown(wait=True)

        for ds, ef in eval_futures:
            try:
                ev = ef.result()
                progress[llm][ds]["__eval__"] = {
                    "status": ev["status"], "elapsed": ev.get("elapsed"),
                    "summary": str(ev.get("summary")),
                }
            except Exception as e:
                progress[llm][ds]["__eval__"] = {"status": f"exception: {e!r}"}
            save_progress()
        eval_ex.shutdown(wait=True)

    for llm in llms:
        if args.serial_datasets:
            _run_serial_dataset_loop(llm)
        else:
            _run_parallel_dataset_loop(llm)

    build_tables(RESULT_ROOT / "tables")
    banner("ALL DONE")
    print(f"Tables: {RESULT_ROOT / 'tables'}")
    print(f"Progress log: {progress_file}")


if __name__ == "__main__":
    main()
