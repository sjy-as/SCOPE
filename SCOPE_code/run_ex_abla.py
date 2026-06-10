"""
Ablation experiment execution framework: N LLMs × 3 datasets × multiple ablation variants.

Ablation models:
    new_modl_wo_semlist_atomr     — new_model with AtomR source routing, no semantic-list
    new_modl_wo_semlist_deepsieve — new_model with DeepSieve source routing, no semantic-list
    new_modl_wo_decomp            — new_model without question decomposition
    new_modl_wo_fallback          — new_model without fallback retry
    new_modl_wo_semlist_plan      — new_model without semantic-list metadata in operator planning
    new_modl_wo_opplan            — new_model without operator planning

These compare against the full new_model (whose predictions are read from the main
experiment output directory). Run run_ex_main.py first to produce new_model predictions.

Layout (ablation experiment owns its own output tree):
    /root/autodl-tmp/eval/result/ex_abla/
        <dataset_slug>/<llm_slug>/
            new_modl_wo_semlist_atomr/predictions.jsonl
            new_modl_wo_semlist_deepsieve/predictions.jsonl
            eval_summary/summary.json   ← only the ablation variants are judged here
        tables/
            abla_tables.json + abla_tables.xlsx   (new_model row reads from
                                                   ex_main/.../eval_summary/summary.json)

Model lists for both experiments live in /root/autodl-tmp/eval/models_config.py.

Recommended invocation (run yourself):

    cd /root/autodl-tmp && python3 run_ex_abla.py
    cd /root/autodl-tmp && python3 run_ex_abla.py --only-llm deepseek-chat
    cd /root/autodl-tmp && python3 run_ex_abla.py --tables-only
"""
from __future__ import annotations

import argparse
import json
import os
import shlex
import subprocess
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

sys.path.insert(0, str(Path(__file__).resolve().parent / "eval"))
from models_config import (  # noqa: E402
    ABLATION_EXPERIMENT_MODELS,
    ABLATION_TABLE_MODELS,
    MODEL_DISPLAY_NAMES,
    MODEL_EVAL_FLAGS,
    MODEL_TO_SOURCE,
    predictions_path,
)

try:
    from openpyxl import Workbook as _XlWorkbook
    from openpyxl.styles import Font as _XlFont, Alignment as _XlAlignment
    from openpyxl.styles import Border as _XlBorder, Side as _XlSide
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


# ---------------------------------------------------------------------------
# Static configuration
# ---------------------------------------------------------------------------

QA_BENCH = Path("/root/autodl-tmp/new_model/qa_bench")
EVAL_DIR = Path("/root/autodl-tmp/eval")
RESULT_ROOT = EVAL_DIR / "result" / "ex_abla"
# new_model (full version) predictions are read from the main experiment output
# directory — this runner only executes the ablation variants.
RESULT_ROOT_MAIN = EVAL_DIR / "result" / "ex_main"
LOG_ROOT = RESULT_ROOT / "_logs"

DATASETS: Dict[str, Dict[str, Any]] = {
    "kg_doc":    {"file": QA_BENCH / "kg-doc-1154.jsonl",   "kb": "kg,doc",    "evaluator": "evaluate_answer_kg_doc.py"},
    "kg_table":  {"file": QA_BENCH / "kg-table-1147.jsonl", "kb": "kg,table",  "evaluator": "evaluate_answer_kg_table.py"},
    "table_doc": {"file": QA_BENCH / "table-doc-1120.jsonl","kb": "table,doc", "evaluator": "evaluate_answer_table_doc.py"},
}

DATASET_ORDER = ["kg_doc", "kg_table", "table_doc"]
# Models that get RUN by this script. Edit eval/models_config.py to extend.
ABLATION_MODELS = list(ABLATION_EXPERIMENT_MODELS)
LLM_ORDER = ["gpt-4o-mini", "gpt-4o", "qwen3-32b", "deepseek-v3", "deepseek-chat"]

# Display names: shared dict from eval/models_config.py, with "new_model" overridden
# to "New Model (Full)" so the ablation table reads naturally.
ABLATION_DISPLAY_NAMES: Dict[str, str] = {**MODEL_DISPLAY_NAMES, "new_model": "New Model (Full)"}

LLM_DISPLAY_NAMES: Dict[str, str] = {
    "gpt-4o-mini":                "GPT-4o-mini",
    "gpt-4o":                     "GPT-4o",
    "qwen3-32b":                  "Qwen3-32B",
    "deepseek-v3":                "DeepSeek-V3",
    "deepseek-chat":              "DeepSeek-Chat",
}

# API keys for ablation models per LLM.
LLM_CONFIGS: Dict[str, Dict[str, Any]] = {
    "gpt-4o-mini": {
        "model": "gpt-4o-mini",
        "base_url": "https://.../v1",
        "keys": {
            "new_modl_wo_semlist_atomr":     "sk-......",
            "new_modl_wo_semlist_deepsieve": "sk-......",
            "new_modl_wo_semlist_plan":      "sk-......",
        },
    },
    "gpt-4o": {
        "model": "gpt-4o",
        "base_url": "https://.../v1",
        "keys": {
            "new_modl_wo_semlist_atomr":     "sk-......",
            "new_modl_wo_semlist_deepsieve": "sk-......",
            "new_modl_wo_semlist_plan":      "sk-......",
        },
    },
    "qwen3-32b": {
        "model": "Qwen3-32B",
        "base_url": "https://.../v1",
        "keys": {
            "new_modl_wo_semlist_atomr":     "sk-......",
            "new_modl_wo_semlist_deepsieve": "sk-......",
            "new_modl_wo_semlist_plan":      "sk-......",
        },
    },
    "deepseek-v3": {
        "model": "DeepSeek-V3",
        "base_url": "https://.../v1",
        "keys": {
            "new_modl_wo_semlist_atomr":     "sk-......",
            "new_modl_wo_semlist_deepsieve": "sk-......",
            "new_modl_wo_semlist_plan":      "sk-......",
        },
    },
    "deepseek-chat": {
        "model": "deepseek-chat",
        "base_url": "https://.../v1",
        "keys": {
            "new_modl_wo_semlist_atomr":     "sk-......",
            "new_modl_wo_semlist_deepsieve": "sk-......",
            "new_modl_wo_decomp":            "sk-......",
            "new_modl_wo_fallback":          "sk-......",
            "new_modl_wo_semlist_plan":      "sk-......",
            "new_modl_wo_opplan":            "sk-......",
        },
    },
    "chatanywhere_deepseek-chat": {
        "model": "deepseek-chat",
        "base_url": "https://.../v1",
        "keys": {
            "new_modl_wo_semlist_atomr":     "sk-......",
            "new_modl_wo_semlist_deepsieve": "sk-......",
            "new_modl_wo_semlist_plan":      "sk-......",
        },
    },
}

# Evaluator (LLM-judge) — fixed across all runs.
JUDGE_LLM_MODEL = "DeepSeek-V4-Flash"
JUDGE_LLM_URL   = "https://.../v1"
JUDGE_API_KEY   = "sk-......"

PER_MODEL_WORKERS = 24
PER_LLM_WORKERS: Dict[str, int] = {
    "deepseek-chat": 24,
}


def workers_for(llm: str) -> int:
    return PER_LLM_WORKERS.get(llm, PER_MODEL_WORKERS)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_name(s: str) -> str:
    return s.replace("/", "_").replace(" ", "_")


def out_root(dataset: str, llm: str) -> Path:
    return RESULT_ROOT / dataset / _safe_name(llm)


def model_out_dir(dataset: str, llm: str, model: str) -> Path:
    return out_root(dataset, llm) / model


def log_path(dataset: str, llm: str, model: str) -> Path:
    LOG_ROOT.mkdir(parents=True, exist_ok=True)
    return LOG_ROOT / f"{dataset}__{_safe_name(llm)}__{model}.log"


def eval_log_path(dataset: str, llm: str) -> Path:
    LOG_ROOT.mkdir(parents=True, exist_ok=True)
    return LOG_ROOT / f"{dataset}__{_safe_name(llm)}__abla_eval.log"


def banner(msg: str) -> None:
    bar = "=" * 78
    print(f"\n{bar}\n{msg}\n{bar}", flush=True)


def file_has_content(p: Path) -> bool:
    return p.exists() and p.is_file() and p.stat().st_size > 0


def run_cmd(
    cmd: List[str],
    log_file: Path,
    cwd: Optional[Path] = None,
    env: Optional[Dict[str, str]] = None,
    timeout: Optional[int] = None,
    quiet: bool = False,
) -> Tuple[int, float]:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    print(f"[exec] {' '.join(shlex.quote(c) for c in cmd)}", flush=True)
    print(f"[exec] cwd={cwd}  log={log_file}", flush=True)
    t0 = time.time()
    with log_file.open("w", encoding="utf-8") as lf:
        lf.write(f"# CMD: {' '.join(shlex.quote(c) for c in cmd)}\n")
        if cwd:
            lf.write(f"# CWD: {cwd}\n")
        lf.flush()
        try:
            proc = subprocess.Popen(
                cmd, cwd=str(cwd) if cwd else None,
                env={**os.environ, **(env or {})},
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                bufsize=1, universal_newlines=True,
                encoding="utf-8", errors="replace",
            )
        except FileNotFoundError as e:
            lf.write(f"\n[ERROR] {e}\n")
            return 127, time.time() - t0

        assert proc.stdout is not None
        for line in proc.stdout:
            if not quiet:
                sys.stdout.write(line)
                sys.stdout.flush()
            lf.write(line)
        proc.wait(timeout=timeout)
    elapsed = time.time() - t0
    return proc.returncode or 0, elapsed


# ---------------------------------------------------------------------------
# Per-model command builders
# ---------------------------------------------------------------------------

def cmd_new_modl_wo_semlist_atomr(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    """new_model ablation: AtomR source routing + no semantic-list content in parse/plan."""
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        "python3", "/root/autodl-tmp/baseline/new_modl_wo_semlist_atomr/run.py",
        "--input", str(ds["file"]),
        "--gold",  str(ds["file"]),
        "--output-dir", str(out_dir),
        "--kb", ds["kb"],
        "--workers", str(PER_MODEL_WORKERS),
        "--api-key",  cfg["keys"]["new_modl_wo_semlist_atomr"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
    ]
    return cmd, Path("/root/autodl-tmp/new_model"), {}


def cmd_new_modl_wo_semlist_deepsieve(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    """new_model ablation: DeepSieve source routing + no semantic-list content in parse/plan."""
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        "python3", "/root/autodl-tmp/baseline/new_modl_wo_semlist_deepsieve/run.py",
        "--input", str(ds["file"]),
        "--gold",  str(ds["file"]),
        "--output-dir", str(out_dir),
        "--kb", ds["kb"],
        "--workers", str(PER_MODEL_WORKERS),
        "--api-key",  cfg["keys"]["new_modl_wo_semlist_deepsieve"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
    ]
    return cmd, Path("/root/autodl-tmp/new_model"), {}


def cmd_new_modl_wo_decomp(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    """new_model ablation: no question decomposition."""
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        "python3", "/root/autodl-tmp/baseline/new_modl_wo_decomp/run.py",
        "--input", str(ds["file"]),
        "--gold",  str(ds["file"]),
        "--output-dir", str(out_dir),
        "--kb", ds["kb"],
        "--workers", str(workers_for(llm)),
        "--api-key",  cfg["keys"]["new_modl_wo_decomp"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
    ]
    return cmd, Path("/root/autodl-tmp/new_model"), {}


def cmd_new_modl_wo_fallback(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    """new_model ablation: no fallback-source retry."""
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        "python3", "/root/autodl-tmp/baseline/new_modl_wo_fallback/run.py",
        "--input", str(ds["file"]),
        "--gold",  str(ds["file"]),
        "--output-dir", str(out_dir),
        "--kb", ds["kb"],
        "--workers", str(workers_for(llm)),
        "--api-key",  cfg["keys"]["new_modl_wo_fallback"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
    ]
    return cmd, Path("/root/autodl-tmp/new_model"), {}


def cmd_new_modl_wo_semlist_plan(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    """new_model ablation: keep planner on, but remove semantic-list metadata from plan generation."""
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        "python3", "/root/autodl-tmp/baseline/new_modl_wo_semlist_plan/run.py",
        "--input", str(ds["file"]),
        "--gold",  str(ds["file"]),
        "--output-dir", str(out_dir),
        "--kb", ds["kb"],
        "--workers", str(workers_for(llm)),
        "--api-key",  cfg["keys"]["new_modl_wo_semlist_plan"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
    ]
    return cmd, Path("/root/autodl-tmp/new_model"), {}


def cmd_new_modl_wo_opplan(dataset: str, llm: str, out_dir: Path) -> Tuple[List[str], Path, Dict[str, str]]:
    """new_model ablation: no operator-tree planning (naive Search step)."""
    cfg = LLM_CONFIGS[llm]
    ds = DATASETS[dataset]
    cmd = [
        "python3", "/root/autodl-tmp/baseline/new_modl_wo_opplan/run.py",
        "--input", str(ds["file"]),
        "--gold",  str(ds["file"]),
        "--output-dir", str(out_dir),
        "--kb", ds["kb"],
        "--workers", str(workers_for(llm)),
        "--api-key",  cfg["keys"]["new_modl_wo_opplan"],
        "--llm-url",  cfg["base_url"],
        "--llm-model", cfg["model"],
    ]
    return cmd, Path("/root/autodl-tmp/new_model"), {}


MODEL_CMD_BUILDERS = {
    "new_modl_wo_semlist_atomr":     cmd_new_modl_wo_semlist_atomr,
    "new_modl_wo_semlist_deepsieve": cmd_new_modl_wo_semlist_deepsieve,
    "new_modl_wo_decomp":        cmd_new_modl_wo_decomp,
    "new_modl_wo_fallback":      cmd_new_modl_wo_fallback,
    "new_modl_wo_semlist_plan":  cmd_new_modl_wo_semlist_plan,
    "new_modl_wo_opplan":        cmd_new_modl_wo_opplan,
}


# ---------------------------------------------------------------------------
# Completion sentinels
# ---------------------------------------------------------------------------

@lru_cache(maxsize=None)
def dataset_size(dataset: str) -> int:
    gold = DATASETS[dataset]["file"]
    n = 0
    with gold.open("rb") as f:
        for _ in f:
            n += 1
    return n


def _count_jsonl_lines(p: Path) -> int:
    if not p.exists() or not p.is_file():
        return 0
    n = 0
    with p.open("rb") as f:
        for _ in f:
            n += 1
    return n


def model_output_progress(dataset: str, llm: str, model: str) -> Tuple[int, int]:
    """Return (done_count, total). Complete iff done >= total > 0."""
    d = model_out_dir(dataset, llm, model)
    total = dataset_size(dataset)
    # Both ablation wrappers delegate to new_model/run.py and produce predictions.jsonl.
    return _count_jsonl_lines(d / "predictions.jsonl"), total


def model_output_exists(dataset: str, llm: str, model: str) -> bool:
    done, total = model_output_progress(dataset, llm, model)
    return done >= total > 0


# ---------------------------------------------------------------------------
# Runners
# ---------------------------------------------------------------------------

def run_one_model(dataset: str, llm: str, model: str, force: bool, quiet: bool = False) -> Dict[str, Any]:
    out_dir = model_out_dir(dataset, llm, model)
    out_dir.mkdir(parents=True, exist_ok=True)

    done, total = model_output_progress(dataset, llm, model)
    if not force and done >= total > 0:
        banner(f"[skip] {dataset} / {llm} / {model}  ({done}/{total} done)")
        return {"status": "skipped", "elapsed": 0.0}

    if done > 0:
        banner(f"[resume] {dataset} / {llm} / {model}  ({done}/{total} done, continuing)")
    else:
        banner(f"[run]  {dataset} / {llm} / {model}  (0/{total})")
    cmd, cwd, env = MODEL_CMD_BUILDERS[model](dataset, llm, out_dir)
    rc, elapsed = run_cmd(cmd, log_path(dataset, llm, model), cwd=cwd, env=env, quiet=quiet)
    status = "ok" if rc == 0 else f"failed(rc={rc})"
    print(f"[done] {dataset}/{llm}/{model}: {status} in {elapsed:.0f}s", flush=True)
    return {"status": status, "elapsed": elapsed, "returncode": rc}


def run_eval(dataset: str, llm: str, force: bool) -> Dict[str, Any]:
    """Re-invoke the evaluator with ablation flags only.

    The full new_model (used as the reference row in the ablation table) is read
    later from the MAIN experiment's summary.json, so we do not pass --new-model
    here — only the ablation variants get judged against the gold set.
    """
    ds = DATASETS[dataset]
    eval_script = EVAL_DIR / ds["evaluator"]
    base = out_root(dataset, llm)
    eval_out_dir = base / "eval_summary"
    summary_file = eval_out_dir / "summary.json"

    if not force and file_has_content(summary_file):
        banner(f"[skip-eval] {dataset} / {llm}  (summary.json present)")
        return {"status": "skipped", "summary": summary_file}

    banner(f"[eval] {dataset} / {llm}  (ablation)")
    eval_out_dir.mkdir(parents=True, exist_ok=True)

    cmd: List[str] = [
        "python3", str(eval_script),
        "--gold", str(ds["file"]),
        "--out-dir", str(eval_out_dir),
        "--llm-url", JUDGE_LLM_URL,
        "--llm-model", JUDGE_LLM_MODEL,
        "--api-key", JUDGE_API_KEY,
        "--max-workers", str(PER_MODEL_WORKERS),
    ]
    for slug in ABLATION_MODELS:
        flag = MODEL_EVAL_FLAGS[slug]
        p = predictions_path(base, dataset, slug)
        if file_has_content(p):
            cmd += [flag.flag, str(p)]

    rc, elapsed = run_cmd(cmd, eval_log_path(dataset, llm))
    status = "ok" if rc == 0 else f"failed(rc={rc})"
    print(f"[eval-done] {dataset}/{llm}: {status} in {elapsed:.0f}s", flush=True)
    return {"status": status, "elapsed": elapsed, "summary": summary_file}


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------

def _read_summary(path: Path) -> Optional[Dict[str, Any]]:
    if not file_has_content(path):
        return None
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"[warn] failed to read {path}: {e}", flush=True)
        return None


# All models shown in the ablation table — comes from eval/models_config.py.
ABLA_MODEL_ORDER = list(ABLATION_TABLE_MODELS)


def _summary_path_for(dataset: str, llm: str, model: str) -> Path:
    """new_model summary lives in the MAIN experiment dir; ablation summaries
    live in this script's ex_abla dir."""
    root = RESULT_ROOT_MAIN if model == "new_model" else RESULT_ROOT
    return root / dataset / _safe_name(llm) / "eval_summary" / "summary.json"


def _extract_cell(summary: Dict[str, Any], model: str) -> Tuple[Optional[float], Optional[float]]:
    """Returns (strict, loose) accuracy on stage sq2."""
    if not summary:
        return None, None
    src = MODEL_TO_SOURCE.get(model, model)
    entry = (summary.get("summaries") or {}).get(src)
    if not entry:
        return None, None
    sq2 = entry.get("sq2")
    if not sq2:
        return None, None
    return sq2.get("strict"), sq2.get("loose")


def build_tables(out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    overall: Dict[str, Any] = {}
    for dataset in DATASET_ORDER:
        rows: List[List[str]] = []
        header = ["model \\ LLM"] + LLM_ORDER
        rows.append(header)
        for model in ABLA_MODEL_ORDER:
            row = [model]
            for llm in LLM_ORDER:
                summary = _read_summary(_summary_path_for(dataset, llm, model))
                strict, loose = _extract_cell(summary or {}, model)
                if strict is None and loose is None:
                    row.append("-")
                else:
                    s = "-" if strict is None else f"{strict:.3f}"
                    l = "-" if loose  is None else f"{loose:.3f}"
                    row.append(f"{s} / {l}")
            rows.append(row)
        overall[dataset] = {"rows": rows}

    with (out_dir / "abla_tables.json").open("w", encoding="utf-8") as f:
        json.dump(overall, f, ensure_ascii=False, indent=2)

    build_excel_table(overall, out_dir)


def build_excel_table(all_tables: Dict[str, Any], out_dir: Path) -> None:
    """Write abla_tables.xlsx.

    Layout: one sheet per LLM that has any data.
    Each sheet:
      - Row 1 : merged title
      - Row 2 : column headers  (Model | KG-Doc | KG-Table | Table-Doc)
      - Row 3 : New Model (Full) — absolute scores, bold, gray background
      - Rows 4+: ablation variants — two-line cell:
                    line 1: strict / loose  (black, bold)
                    line 2: (Δstrict / Δloose) vs Full Model
                             red  if Δstrict < 0  (worse)
                             green if Δstrict > 0  (better)
                             gray  if no change
    """
    if not _OPENPYXL:
        print("[tables] openpyxl not installed; skipping Excel output (pip install openpyxl)")
        return

    DATASET_LABELS = {"kg_doc": "KG-Doc", "kg_table": "KG-Table", "table_doc": "Table-Doc"}

    def _val(ds: str, model: str, llm: str) -> str:
        rows = all_tables.get(ds, {}).get("rows", [])
        hdr = rows[0] if rows else []
        if llm not in hdr:
            return "-"
        col = hdr.index(llm)
        for row in rows[1:]:
            if row[0] == model:
                return row[col] if col < len(row) else "-"
        return "-"

    def _parse(v: str) -> Optional[Tuple[float, float]]:
        if v == "-":
            return None
        parts = v.split(" / ")
        if len(parts) != 2:
            return None
        try:
            return float(parts[0]), float(parts[1])
        except ValueError:
            return None

    def _thin():   return _XlSide(style="thin")
    def _medium(): return _XlSide(style="medium")
    def _brd(top=None, bottom=None, left=None, right=None):
        return _XlBorder(top=top, bottom=bottom, left=left, right=right)

    wb = _XlWorkbook()
    wb.remove(wb.active)

    for llm in LLM_ORDER:
        # Skip LLMs with no data at all.
        if all(_val(ds, m, llm) == "-" for ds in DATASET_ORDER for m in ABLA_MODEL_ORDER):
            continue

        sheet_name = LLM_DISPLAY_NAMES.get(llm, llm)[:31]
        ws = wb.create_sheet(title=sheet_name)

        ws.column_dimensions["A"].width = 28
        for col_letter in ["B", "C", "D"]:
            ws.column_dimensions[col_letter].width = 24

        # ── Row 1: title ──────────────────────────────────────────────────────
        ws.merge_cells("A1:D1")
        tc = ws["A1"]
        tc.value = (f"Ablation Results  ·  {LLM_DISPLAY_NAMES.get(llm, llm)}"
                    f"  (strict / loose accuracy,  Δ = ablation − full model)")
        tc.font      = _XlFont(name="Arial", bold=True, size=11)
        tc.alignment = _XlAlignment(horizontal="center", vertical="center")
        tc.border    = _brd(_medium(), _medium(), _medium(), _medium())
        ws.row_dimensions[1].height = 24

        # ── Row 2: column headers ─────────────────────────────────────────────
        headers = ["Model"] + [DATASET_LABELS[ds] for ds in DATASET_ORDER]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font      = _XlFont(name="Arial", bold=True, size=10)
            cell.alignment = _XlAlignment(horizontal="center", vertical="center")
            left_brd  = _medium() if ci == 1 else _thin()
            right_brd = _medium() if ci == 4 else _thin()
            cell.border = _brd(_medium(), _medium(), left_brd, right_brd)
        ws.row_dimensions[2].height = 18

        # ── Pre-compute full-model scores for delta calculation ───────────────
        full: Dict[str, Optional[Tuple[float, float]]] = {
            ds: _parse(_val(ds, "new_model", llm)) for ds in DATASET_ORDER
        }

        # ── Data rows ─────────────────────────────────────────────────────────
        n_models = len(ABLA_MODEL_ORDER)
        for row_i, model in enumerate(ABLA_MODEL_ORDER):
            cur      = 3 + row_i
            is_last  = (row_i == n_models - 1)
            is_full  = (model == "new_model")
            bot      = _medium() if is_last else _thin()
            ws.row_dimensions[cur].height = 16 if is_full else 30

            # Model name cell
            mc = ws.cell(row=cur, column=1,
                         value=ABLATION_DISPLAY_NAMES.get(model, model))
            mc.font      = _XlFont(name="Arial", size=10, bold=is_full)
            mc.alignment = _XlAlignment(horizontal="left", vertical="center",
                                        indent=1, wrap_text=False)
            mc.border    = _brd(_thin(), bot, _medium(), _thin())

            # Score cells (one per dataset)
            for col_idx, ds in enumerate(DATASET_ORDER, start=2):
                right_brd = _medium() if col_idx == 4 else _thin()
                cell = ws.cell(row=cur, column=col_idx)
                cell.border = _brd(_thin(), bot, _thin(), right_brd)

                v      = _val(ds, model, llm)
                parsed = _parse(v)

                if parsed is None:
                    # No data
                    cell.value     = "—"
                    cell.font      = _XlFont(name="Arial", size=10)
                    cell.alignment = _XlAlignment(horizontal="center", vertical="center")
                elif is_full:
                    # Full model: just the score, bold
                    cell.value     = f"{parsed[0]:.3f} / {parsed[1]:.3f}"
                    cell.font      = _XlFont(name="Arial", size=10, bold=True)
                    cell.alignment = _XlAlignment(horizontal="center", vertical="center")
                else:
                    # Ablation row: score + delta on second line
                    full_p = full.get(ds)
                    if full_p:
                        d_s = parsed[0] - full_p[0]
                        d_l = parsed[1] - full_p[1]
                        sign_s = "+" if d_s >= 0 else ""
                        sign_l = "+" if d_l >= 0 else ""
                        delta_str = f"({sign_s}{d_s:.3f} / {sign_l}{d_l:.3f})"
                        cell.value = f"{parsed[0]:.3f} / {parsed[1]:.3f}\n{delta_str}"
                        cell.font  = _XlFont(name="Arial", size=10)
                    else:
                        cell.value = f"{parsed[0]:.3f} / {parsed[1]:.3f}"
                        cell.font  = _XlFont(name="Arial", size=10)
                    cell.alignment = _XlAlignment(horizontal="center", vertical="center",
                                                  wrap_text=True)

        ws.freeze_panes = "B3"

    xlsx_path = out_dir / "abla_tables.xlsx"
    wb.save(xlsx_path)
    print(f"[tables] wrote {xlsx_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description="Ablation experiment: N LLMs × 3 datasets × multiple ablation variants")
    ap.add_argument("--only-llm",     choices=LLM_ORDER,       help="Restrict to this LLM")
    ap.add_argument("--only-dataset", choices=DATASET_ORDER,   help="Restrict to this dataset")
    ap.add_argument("--only-model",   choices=ABLATION_MODELS, help="Restrict to this ablation model")
    ap.add_argument("--skip-eval",   action="store_true", help="Don't run evaluators")
    ap.add_argument("--skip-runs",   action="store_true", help="Don't run any model — eval / tables only")
    ap.add_argument("--tables-only", action="store_true", help="Just aggregate existing summaries into tables")
    ap.add_argument("--force",       action="store_true", help="Re-run even if outputs exist")
    ap.add_argument("--serial-models", action="store_true",
                    help="Run ablation models serially within each (llm, dataset). "
                         "Default: run them in parallel.")
    ap.add_argument("--serial-datasets", action="store_true",
                    help="Process datasets one at a time.")
    args = ap.parse_args()

    RESULT_ROOT.mkdir(parents=True, exist_ok=True)
    LOG_ROOT.mkdir(parents=True, exist_ok=True)

    if args.tables_only:
        build_tables(RESULT_ROOT / "tables")
        return

    llms     = [args.only_llm]     if args.only_llm     else LLM_ORDER
    datasets = [args.only_dataset] if args.only_dataset else DATASET_ORDER
    models   = [args.only_model]   if args.only_model   else ABLATION_MODELS

    progress: Dict[str, Any] = {}
    progress_file = RESULT_ROOT / "_abla_progress.json"
    if progress_file.exists():
        try:
            progress = json.loads(progress_file.read_text(encoding="utf-8"))
        except Exception:
            progress = {}

    def save_progress():
        progress_file.write_text(
            json.dumps(progress, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _run_serial_dataset_loop(llm: str) -> None:
        for dataset in datasets:
            banner(f"### ABLA  LLM={llm}  DATASET={dataset}")
            progress.setdefault(llm, {}).setdefault(dataset, {})

            if not args.skip_runs:
                if args.serial_models or len(models) <= 1:
                    for model in models:
                        res = run_one_model(dataset, llm, model, force=args.force)
                        progress[llm][dataset][model] = res
                        save_progress()
                else:
                    banner(f"[parallel] launching {len(models)} ablation models concurrently "
                           f"for {dataset} / {llm}")
                    with ThreadPoolExecutor(max_workers=len(models)) as ex:
                        futs = {ex.submit(run_one_model, dataset, llm, m, args.force, quiet=True): m
                                for m in models}
                        for fut in as_completed(futs):
                            m = futs[fut]
                            try:
                                res = fut.result()
                            except Exception as e:
                                res = {"status": f"exception: {e!r}", "elapsed": 0.0}
                            progress[llm][dataset][m] = res
                            save_progress()

            if not args.skip_eval:
                ev = run_eval(dataset, llm, force=args.force)
                progress[llm][dataset]["__eval__"] = {
                    "status":  ev["status"], "elapsed": ev.get("elapsed"),
                    "summary": str(ev.get("summary")),
                }
                save_progress()

    def _run_parallel_dataset_loop(llm: str) -> None:
        """Default: launch every (dataset, model) for this LLM concurrently."""
        for dataset in datasets:
            progress.setdefault(llm, {}).setdefault(dataset, {})

        if args.skip_runs:
            if not args.skip_eval:
                with ThreadPoolExecutor(max_workers=len(datasets)) as eval_ex:
                    futs = {eval_ex.submit(run_eval, ds, llm, args.force): ds for ds in datasets}
                    for fut in as_completed(futs):
                        ds = futs[fut]
                        try:
                            ev = fut.result()
                        except Exception as e:
                            ev = {"status": f"exception: {e!r}"}
                        progress[llm][ds]["__eval__"] = {
                            "status": ev["status"], "elapsed": ev.get("elapsed"),
                            "summary": str(ev.get("summary")),
                        }
                        save_progress()
            return

        remaining = {ds: set(models) for ds in datasets}
        state_lock = threading.Lock()
        eval_futures: List[Tuple[str, Any]] = []
        n_model_slots = len(datasets) * len(models)

        banner(f"### ABLA  LLM={llm}  —  launching {len(datasets)} datasets × {len(models)} "
               f"ablation models  ({n_model_slots} concurrent; logs in {LOG_ROOT}/)")

        model_ex = ThreadPoolExecutor(max_workers=n_model_slots)
        eval_ex = ThreadPoolExecutor(max_workers=len(datasets))
        try:
            model_futs = {}
            for ds in datasets:
                for m in models:
                    f = model_ex.submit(run_one_model, ds, llm, m, args.force, quiet=True)
                    model_futs[f] = (ds, m)

            for fut in as_completed(model_futs):
                ds, m = model_futs[fut]
                try:
                    res = fut.result()
                except Exception as e:
                    res = {"status": f"exception: {e!r}", "elapsed": 0.0}
                fire_eval = False
                with state_lock:
                    progress[llm][ds][m] = res
                    save_progress()
                    remaining[ds].discard(m)
                    if not remaining[ds]:
                        fire_eval = True
                if fire_eval and not args.skip_eval:
                    print(f"[eval-trigger] {ds}/{llm}: all ablation models done, "
                          f"launching evaluator", flush=True)
                    eval_futures.append((ds, eval_ex.submit(run_eval, ds, llm, args.force)))
        finally:
            model_ex.shutdown(wait=True)

        for ds, ef in eval_futures:
            try:
                ev = ef.result()
                progress[llm][ds]["__eval__"] = {
                    "status": ev["status"], "elapsed": ev.get("elapsed"),
                    "summary": str(ev.get("summary")),
                }
            except Exception as e:
                progress[llm][ds]["__eval__"] = {"status": f"exception: {e!r}"}
            save_progress()
        eval_ex.shutdown(wait=True)

    for llm in llms:
        if args.serial_datasets:
            _run_serial_dataset_loop(llm)
        else:
            _run_parallel_dataset_loop(llm)

    build_tables(RESULT_ROOT / "tables")
    banner("ABLATION DONE")
    print(f"Tables: {RESULT_ROOT / 'tables'}")
    print(f"Progress log: {progress_file}")


if __name__ == "__main__":
    main()
